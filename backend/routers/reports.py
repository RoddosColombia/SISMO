"""reports.py — Generic Excel export endpoint (HOTFIX 21.1 FIX #3)."""
import io
from pydantic import BaseModel
from typing import Any, Dict

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
import openpyxl
from openpyxl.styles import Font

from database import db
from dependencies import get_current_user

router = APIRouter(prefix="/reports", tags=["reports"])


class ExcelRequest(BaseModel):
    reportType: str
    filters: Dict[str, Any] = {}


@router.post("/excel")
async def download_excel(
    body: ExcelRequest,
    current_user=Depends(get_current_user),
):
    """Generate and stream an Excel file. Supported reportType: 'loanbooks', 'chat'."""
    wb = openpyxl.Workbook()
    ws = wb.active

    if body.reportType == "loanbooks":
        ws.title = "Loanbooks"
        headers = [
            "Código", "Cliente", "Plan", "Estado", "Valor Moto",
            "Cuotas Pagadas", "Cuotas Totales", "Saldo Pendiente",
            "Fecha Factura", "Moto",
        ]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)

        loans = await db.loanbook.find({}, {"_id": 0}).to_list(2000)
        for loan in loans:
            ws.append([
                loan.get("codigo", ""),
                loan.get("cliente_nombre", ""),
                loan.get("plan", ""),
                loan.get("estado", ""),
                loan.get("valor_moto", 0),
                loan.get("cuotas_pagadas", 0),
                loan.get("num_cuotas", 0),
                loan.get("saldo_pendiente", 0),
                loan.get("fecha_factura", ""),
                loan.get("moto_descripcion", ""),
            ])
        fname = "RODDOS_Loanbooks.xlsx"

    elif body.reportType == "chat":
        ws.title = "Conversaciones"
        headers = ["ID", "Usuario", "Fecha", "Rol", "Contenido"]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)

        messages = await db.chat_messages.find({}, {"_id": 0}).sort("created_at", -1).to_list(1000)
        for msg in messages:
            ws.append([
                msg.get("id", ""),
                msg.get("user_id", ""),
                msg.get("created_at", ""),
                msg.get("role", ""),
                str(msg.get("content", ""))[:500],
            ])
        fname = "RODDOS_Chat.xlsx"

    else:
        raise HTTPException(
            status_code=400,
            detail=f"reportType '{body.reportType}' no soportado. Use 'loanbooks' o 'chat'.",
        )

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )
