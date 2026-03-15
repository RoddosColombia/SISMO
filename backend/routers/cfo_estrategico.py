"""cfo_estrategico.py — BUILD 11: Agente CFO Estratégico RODDOS.

Endpoints:
  GET/POST  /cfo/financiero/config          — gastos_fijos_semanales y parámetros
  GET       /cfo/deudas/plantilla          — descarga plantilla Excel para carga de deudas
  POST      /cfo/deudas/cargar             — parse Excel → clasificación automática (preview)
  POST      /cfo/deudas/confirmar          — guardar deudas confirmadas en MongoDB
  GET       /cfo/deudas                    — listar deudas
  PATCH     /cfo/deudas/{id}              — actualizar/reclasificar deuda
  GET       /cfo/plan-ingresos            — plan semanal desde loanbooks
  GET       /cfo/plan-deudas             — plan avalancha deuda no productiva
  GET       /cfo/reporte-lunes           — reporte CFO para el lunes
  GET       /cfo/cuotas-iniciales        — saldos pendientes cuota inicial
  GET       /cfo/indicadores             — indicadores clave (piso creditos, etc.)
"""
import io
import uuid
import logging
import unicodedata
from datetime import datetime, timezone, date, timedelta
from typing import Optional

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse
from pydantic import BaseModel

from database import db
from dependencies import get_current_user

router = APIRouter(prefix="/cfo", tags=["cfo-estrategico"])
logger = logging.getLogger(__name__)

RECAUDO_SEMANAL_BASE = 1_509_500
TICKET_PROMEDIO      = 167_722

# ── helpers ───────────────────────────────────────────────────────────────────

def _next_wednesday(from_date: date) -> date:
    days = (2 - from_date.weekday()) % 7
    return from_date + timedelta(days=days)


def _clasificar_deuda(acreedor: str, descripcion: str) -> str:
    texto = f"{acreedor} {descripcion}".lower()
    productivas = [
        "auteco", "moto", "inventario", "leasing", "credito auteco",
        "financiacion inventario", "compra motos", "equipo trabajo",
    ]
    no_productivas = [
        "arriendo", "arrendamiento", "tarjeta", "credito personal",
        "prestamo personal", "servicio", "agua", "luz", "gas",
        "internet", "telefono", "impuesto", "iva", "retencion",
        "nomina", "salario", "multa", "consumo",
    ]
    if any(kw in texto for kw in productivas):
        return "productiva"
    if any(kw in texto for kw in no_productivas):
        return "no_productiva"
    return "no_productiva"


def _prioridad(tasa_mensual: float, tipo: str) -> int:
    """Mayor tasa → menor número (mayor prioridad) — método avalancha."""
    if tipo == "no_productiva":
        return max(1, 10 - int(tasa_mensual))
    return 99


def _fmt_cop(n: float) -> str:
    return f"${n:,.0f}"


# ── GET /cfo/financiero/config ─────────────────────────────────────────────────
@router.get("/financiero/config")
async def get_financiero_config(current_user=Depends(get_current_user)):
    cfg = await db.cfo_financiero_config.find_one({}, {"_id": 0})
    return cfg or {
        "gastos_fijos_semanales": 0,
        "reserva_minima_semanas": 2,
        "limite_compromisos_pct": 0.60,
        "ticket_promedio_cuota": TICKET_PROMEDIO,
        "recaudo_semanal_base": RECAUDO_SEMANAL_BASE,
        "objetivo_deuda_np_meses": 3,
        "configurado": False,
    }


class FinancieroConfigReq(BaseModel):
    gastos_fijos_semanales: float
    reserva_minima_semanas: int = 2
    limite_compromisos_pct: float = 0.60
    objetivo_deuda_np_meses: int = 3


@router.post("/financiero/config")
async def save_financiero_config(req: FinancieroConfigReq, current_user=Depends(get_current_user)):
    data = req.model_dump()
    data["ticket_promedio_cuota"] = TICKET_PROMEDIO
    data["recaudo_semanal_base"]  = RECAUDO_SEMANAL_BASE
    data["configurado"] = True
    data["updated_at"]  = datetime.now(timezone.utc).isoformat()
    await db.cfo_financiero_config.update_one({}, {"$set": data}, upsert=True)

    # Recalculate indicadores
    creditos_minimos = -(-data["gastos_fijos_semanales"] // TICKET_PROMEDIO)
    margen = RECAUDO_SEMANAL_BASE - data["gastos_fijos_semanales"] - (data["gastos_fijos_semanales"] * data["reserva_minima_semanas"])
    return {
        "ok": True,
        "config": data,
        "indicadores": {
            "creditos_minimos": int(creditos_minimos),
            "margen_semanal": margen,
        },
    }


# ── GET /cfo/deudas/plantilla ──────────────────────────────────────────────────
HEADER_COLOR = "0F2A5C"
EXAMPLE_ACREEDOR = "Auteco Kawasaki"

_HEADER_COLS = [
    "Acreedor", "Descripcion", "Monto_Total", "Monto_Pagado",
    "Tasa_Mensual_Pct", "Fecha_Vencimiento", "Tipo", "Prioridad",
]
_INSTRUCT_ROWS = [
    ("Acreedor",           "Nombre del banco, persona o empresa a quien se debe", "Auteco Kawasaki",           "✅"),
    ("Descripcion",        "Concepto de la deuda",                                "Financiación inventario",   "✅"),
    ("Monto_Total",        "Valor original de la deuda en pesos, sin puntos ni comas", "45000000",            "✅"),
    ("Monto_Pagado",       "Cuánto se ha pagado hasta hoy. 0 si no se ha pagado nada", "5000000",             "✅"),
    ("Tasa_Mensual_Pct",   "Tasa de interés mensual en porcentaje. 0 si no cobra intereses", "1.5",           "✅"),
    ("Fecha_Vencimiento",  "Fecha límite de pago en formato YYYY-MM-DD",          "2026-06-30",                "✅"),
    ("Tipo",               "productiva = genera ingresos | no_productiva = no genera ingresos", "productiva", "✅"),
    ("Prioridad",          "Orden de pago. 1 = pagar primero",                    "1",                         "Opcional"),
]


@router.get("/deudas/plantilla")
async def descargar_plantilla_deudas(current_user=Depends(get_current_user)):
    """Genera y descarga la plantilla Excel oficial para carga de deudas RODDOS."""
    wb = openpyxl.Workbook()

    # ── Hoja 1: Deudas ────────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Deudas"

    hdr_fill  = PatternFill("solid", fgColor=HEADER_COLOR)
    hdr_font  = Font(bold=True, color="FFFFFF", size=11)
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_side = Side(style="thin", color="CCCCCC")
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    # Header row
    for col_idx, col_name in enumerate(_HEADER_COLS, start=1):
        cell = ws1.cell(row=1, column=col_idx, value=col_name)
        cell.fill   = hdr_fill
        cell.font   = hdr_font
        cell.alignment = hdr_align
        cell.border = thin_border

    ws1.row_dimensions[1].height = 22

    # Example row (row 2) — highlight lightly so user sees format
    example_fill = PatternFill("solid", fgColor="E8F0FE")
    example_vals = [EXAMPLE_ACREEDOR, "Financiación inventario motos", 45000000, 5000000, 1.5, "2026-06-30", "productiva", 1]
    for col_idx, val in enumerate(example_vals, start=1):
        cell = ws1.cell(row=2, column=col_idx, value=val)
        cell.fill   = example_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center" if col_idx > 2 else "left", vertical="center")

    # Column widths
    col_widths = [22, 28, 15, 15, 17, 18, 16, 12]
    for i, w in enumerate(col_widths, start=1):
        ws1.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # Data validation: Tipo dropdown (col G = 7)
    dv_tipo = DataValidation(
        type="list",
        formula1='"productiva,no_productiva"',
        allow_blank=True,
        showDropDown=False,
    )
    dv_tipo.error       = "Usa: productiva o no_productiva"
    dv_tipo.errorTitle  = "Valor inválido"
    dv_tipo.prompt      = "Selecciona el tipo de deuda"
    dv_tipo.promptTitle = "Tipo de deuda"
    ws1.add_data_validation(dv_tipo)
    dv_tipo.sqref = "G2:G1000"

    # ── Hoja 2: Instrucciones ─────────────────────────────────────────────────
    ws2 = wb.create_sheet("Instrucciones")
    ws2.sheet_state = "visible"

    title_cell = ws2.cell(row=1, column=1, value="INSTRUCCIONES — Carga de Deudas RODDOS")
    title_cell.font  = Font(bold=True, color=HEADER_COLOR, size=14)
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws2.row_dimensions[1].height = 28
    ws2.merge_cells("A1:D1")

    ws2.cell(row=3, column=1, value="Campo").font       = Font(bold=True)
    ws2.cell(row=3, column=2, value="Descripción").font = Font(bold=True)
    ws2.cell(row=3, column=3, value="Ejemplo").font     = Font(bold=True)
    ws2.cell(row=3, column=4, value="Obligatorio").font = Font(bold=True)
    for col in range(1, 5):
        c = ws2.cell(row=3, column=col)
        c.fill   = PatternFill("solid", fgColor="D9E1F2")
        c.border = thin_border
        c.alignment = Alignment(horizontal="center", vertical="center")

    for r_idx, (campo, desc, ej, oblig) in enumerate(_INSTRUCT_ROWS, start=4):
        ws2.cell(row=r_idx, column=1, value=campo).font  = Font(bold=True)
        ws2.cell(row=r_idx, column=2, value=desc)
        ws2.cell(row=r_idx, column=3, value=ej).font     = Font(italic=True, color="1F497D")
        ws2.cell(row=r_idx, column=4, value=oblig).alignment = Alignment(horizontal="center")
        for col in range(1, 5):
            ws2.cell(row=r_idx, column=col).border = thin_border

    ws2.column_dimensions["A"].width = 22
    ws2.column_dimensions["B"].width = 50
    ws2.column_dimensions["C"].width = 22
    ws2.column_dimensions["D"].width = 12

    # Notas
    notes_start = len(_INSTRUCT_ROWS) + 6
    notas = [
        "NOTAS IMPORTANTES:",
        "• No modificar los nombres de las columnas (fila 1 en hoja Deudas)",
        "• No eliminar la fila de ejemplo — el sistema la ignora automáticamente",
        "• Una fila por cada deuda",
        "• Guardar como .xlsx antes de subir al sistema",
    ]
    for i, nota in enumerate(notas):
        c = ws2.cell(row=notes_start + i, column=1, value=nota)
        c.alignment = Alignment(wrap_text=True)
        if i == 0:
            c.font = Font(bold=True, color=HEADER_COLOR)
        ws2.merge_cells(f"A{notes_start + i}:D{notes_start + i}")

    # Serialize to bytes
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="RODDOS_Plantilla_Deudas.xlsx"'},
    )


# ── POST /cfo/deudas/cargar ────────────────────────────────────────────────────
ALIASES = {
    "acreedor":          ["acreedor", "nombre", "creditor", "entidad", "banco"],
    "monto_total":       ["monto_total", "monto", "total", "valor", "saldo", "deuda"],
    "monto_pagado":      ["monto_pagado", "pagado", "abonado", "paid"],
    "tasa_mensual":      ["tasa_mensual_pct", "tasa_mensual", "tasa", "interes", "rate", "mensual"],
    "fecha_vencimiento": ["fecha_vencimiento", "vencimiento", "fecha", "due", "vence"],
    "descripcion":       ["descripcion", "descripción", "concepto", "detalle", "notas"],
    "tipo":              ["tipo", "clasificacion", "clasificación", "type", "categoria"],
    "prioridad":         ["prioridad", "priority", "orden"],
}

REQUIRED_FIELDS = ["acreedor", "monto_total"]


def _normalize_header(h: str) -> str:
    """Lower, strip accents and spaces for fuzzy matching."""
    return unicodedata.normalize("NFD", h.strip().lower()).encode("ascii", "ignore").decode()


def _parse_number(raw, field: str, row_num: int) -> tuple[float, str | None]:
    """Parse a monetary/numeric value. Returns (value, error_msg|None)."""
    if raw is None:
        return 0.0, None
    if isinstance(raw, (int, float)):
        return float(raw), None
    # string: remove $, dots used as thousands, commas
    cleaned = str(raw).replace("$", "").replace(".", "").replace(",", ".").strip()
    try:
        return float(cleaned), None
    except ValueError:
        return 0.0, f"Fila {row_num} — {field}: '{raw}' no es un número válido. Usa solo dígitos, sin puntos ni símbolo $"


@router.post("/deudas/cargar")
async def cargar_deudas_excel(
    file: UploadFile = File(...),
    current_user=Depends(get_current_user),
):
    """Recibe un archivo Excel, parsea las columnas y clasifica cada deuda."""
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="El archivo debe ser .xlsx o .xls")

    contents = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
        ws = wb.active
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Error al leer el archivo: {e}")

    # Read header row
    headers_raw = [str(cell.value or "").strip() for cell in ws[1]]
    headers_norm = [_normalize_header(h) for h in headers_raw]

    # Smart column mapping
    col_map: dict[str, int] = {}
    for field, field_aliases in ALIASES.items():
        for i, h in enumerate(headers_norm):
            if any(_normalize_header(a) in h or h in _normalize_header(a) for a in field_aliases):
                col_map[field] = i
                break

    # Check required fields
    missing = [f for f in REQUIRED_FIELDS if f not in col_map]
    if missing:
        return {
            "ok": False,
            "error": (
                f"El archivo no tiene las columnas requeridas. "
                f"Columnas faltantes: {missing}. "
                f"Descarga la plantilla correcta con el botón de arriba."
            ),
            "columnas_detectadas": headers_raw,
            "sugerencia": "Descarga la plantilla oficial y llena tus datos en ella.",
        }

    # Parse rows
    deudas_preview = []
    row_errors: list[str] = []
    first_data_row = True

    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if all(v is None for v in row):
            continue

        def _get(field):
            idx = col_map.get(field)
            return row[idx] if idx is not None and idx < len(row) else None

        acreedor = str(_get("acreedor") or "").strip()
        if not acreedor:
            continue

        # Skip example row (Auteco Kawasaki in first data position)
        if first_data_row and acreedor.lower() == EXAMPLE_ACREEDOR.lower():
            first_data_row = False
            continue
        first_data_row = False

        # Parse numbers with validation
        monto, err_monto = _parse_number(_get("monto_total"), "Monto_Total", row_num)
        pagado, err_pag  = _parse_number(_get("monto_pagado"), "Monto_Pagado", row_num)
        tasa,   err_tasa = _parse_number(_get("tasa_mensual"), "Tasa_Mensual_Pct", row_num)

        for err in [err_monto, err_pag, err_tasa]:
            if err:
                row_errors.append(err)

        desc = str(_get("descripcion") or "").strip()
        venc = _get("fecha_vencimiento")

        if isinstance(venc, datetime):
            venc = venc.date().isoformat()
        elif venc:
            venc = str(venc)[:10]

        # Tipo: from column if present, else auto-classify
        tipo_raw = str(_get("tipo") or "").strip().lower() if col_map.get("tipo") is not None else ""
        if tipo_raw in ("productiva", "no_productiva"):
            tipo = tipo_raw
        else:
            tipo = _clasificar_deuda(acreedor, desc)

        prioridad_raw = _get("prioridad")
        prioridad = int(prioridad_raw) if prioridad_raw and str(prioridad_raw).isdigit() else _prioridad(tasa, tipo)

        saldo = max(0.0, monto - pagado)
        estado = "activa"
        if venc:
            try:
                if date.fromisoformat(venc) < date.today():
                    estado = "vencida"
            except Exception:
                pass

        deudas_preview.append({
            "id":                str(uuid.uuid4()),
            "acreedor":          acreedor,
            "monto_total":       monto,
            "monto_pagado":      pagado,
            "saldo_pendiente":   saldo,
            "tasa_mensual":      tasa,
            "fecha_vencimiento": venc,
            "descripcion":       desc,
            "tipo":              tipo,
            "prioridad_pago":    prioridad,
            "estado":            estado,
        })

    if not deudas_preview and not row_errors:
        return {
            "ok": False,
            "error": "El archivo no tiene filas con datos. Llena al menos una deuda y vuelve a subir.",
        }

    # Compute summary
    total_prod = sum(d["monto_total"] for d in deudas_preview if d["tipo"] == "productiva")
    total_np   = sum(d["monto_total"] for d in deudas_preview if d["tipo"] == "no_productiva")
    cfg = await db.cfo_financiero_config.find_one({}, {"_id": 0}) or {}
    gastos = cfg.get("gastos_fijos_semanales", 0)

    carga_mensual = sum(
        d["monto_total"] * (d["tasa_mensual"] / 100)
        for d in deudas_preview if d["tipo"] == "no_productiva"
    )
    recaudo_mensual  = RECAUDO_SEMANAL_BASE * 4
    pct_comprometido = (((gastos * 4) + carga_mensual) / recaudo_mensual * 100) if recaudo_mensual else 0
    meses_libres     = round(total_np / max(1, RECAUDO_SEMANAL_BASE * 4 - gastos * 4), 1) if total_np > 0 else 0

    return {
        "ok": True,
        "deudas":  deudas_preview,
        "advertencias": row_errors,
        "resumen": {
            "total_productiva":         total_prod,
            "total_no_productiva":      total_np,
            "carga_financiera_mensual": carga_mensual,
            "pct_recaudo_comprometido": round(pct_comprometido, 1),
            "meses_para_liberar_np":    meses_libres,
        },
        "columnas_mapeadas": col_map,
    }


# ── POST /cfo/deudas/confirmar ─────────────────────────────────────────────────
class DeudaItem(BaseModel):
    id: str
    acreedor: str
    monto_total: float
    monto_pagado: float = 0
    saldo_pendiente: float
    tasa_mensual: float
    fecha_vencimiento: Optional[str] = None
    descripcion: str = ""
    tipo: str  # "productiva" | "no_productiva"
    prioridad_pago: int
    estado: str = "activa"


class ConfirmarDeudasReq(BaseModel):
    deudas: list[DeudaItem]


@router.post("/deudas/confirmar")
async def confirmar_deudas(req: ConfirmarDeudasReq, current_user=Depends(get_current_user)):
    now = datetime.now(timezone.utc).isoformat()
    docs = []
    for d in req.deudas:
        doc = d.model_dump()
        doc["fecha_carga"] = now
        doc["cargado_por"] = current_user.get("email", "")
        docs.append(doc)

    if docs:
        # Clear existing and replace
        await db.cfo_deudas.delete_many({})
        await db.cfo_deudas.insert_many(docs)

    return {"ok": True, "guardadas": len(docs)}


# ── GET /cfo/deudas ────────────────────────────────────────────────────────────
@router.get("/deudas")
async def get_deudas(current_user=Depends(get_current_user)):
    deudas = await db.cfo_deudas.find({}, {"_id": 0}).sort("prioridad_pago", 1).to_list(100)
    return deudas


# ── PATCH /cfo/deudas/{id} ─────────────────────────────────────────────────────
class DeudaUpdate(BaseModel):
    tipo: Optional[str] = None
    estado: Optional[str] = None
    monto_pagado: Optional[float] = None
    saldo_pendiente: Optional[float] = None


@router.patch("/deudas/{deuda_id}")
async def update_deuda(deuda_id: str, body: DeudaUpdate, current_user=Depends(get_current_user)):
    updates = {k: v for k, v in body.model_dump().items() if v is not None}
    if not updates:
        raise HTTPException(status_code=400, detail="Nada que actualizar")
    if "tipo" in updates:
        updates["prioridad_pago"] = _prioridad(0, updates["tipo"])
    updates["updated_at"] = datetime.now(timezone.utc).isoformat()
    await db.cfo_deudas.update_one({"id": deuda_id}, {"$set": updates})
    return {"ok": True}


# ── GET /cfo/plan-ingresos ─────────────────────────────────────────────────────
@router.get("/plan-ingresos")
async def get_plan_ingresos(semanas: int = 8, current_user=Depends(get_current_user)):
    today = date.today()
    loanbooks = await db.loanbook.find(
        {"estado": "activo", "cuota_valor": {"$gt": 0}}, {"_id": 0}
    ).to_list(100)

    cuotas_iniciales_lbs = await db.loanbook.find(
        {"cuota_inicial_pendiente": {"$gt": 0}}, {"_id": 0}
    ).to_list(100)

    semanas_data = []
    for w in range(semanas):
        ws = today + timedelta(weeks=w)
        we = today + timedelta(weeks=w + 1)
        wed = _next_wednesday(ws)
        if wed >= we:
            wed = _next_wednesday(today + timedelta(weeks=w, days=-6))
            if wed < ws:
                wed = ws + timedelta(days=(2 - ws.weekday()) % 7)

        cuotas_sem = []
        for lb in loanbooks:
            cv = lb.get("cuota_valor", 0) or 0
            for c in lb.get("cuotas", []):
                if c.get("estado") == "pendiente":
                    fv = c.get("fecha_vencimiento", "")
                    if fv:
                        try:
                            d = date.fromisoformat(fv[:10])
                            if ws <= d < we:
                                cuotas_sem.append({
                                    "cliente": lb.get("cliente_nombre", ""),
                                    "codigo":  lb.get("codigo", ""),
                                    "numero":  c.get("numero", 0),
                                    "fecha":   fv[:10],
                                    "valor":   cv,
                                })
                        except Exception:
                            pass

        semanas_data.append({
            "semana":          w + 1,
            "miercoles":       wed.isoformat(),
            "recaudo_cartera": sum(c["valor"] for c in cuotas_sem),
            "num_cuotas":      len(cuotas_sem),
            "cuotas":          cuotas_sem,
        })

    cfg = await db.cfo_financiero_config.find_one({}, {"_id": 0}) or {}
    gastos = cfg.get("gastos_fijos_semanales", 0)
    creditos_minimos = int(-(-gastos // TICKET_PROMEDIO)) if gastos > 0 else 0

    return {
        "semanas":                  semanas_data,
        "recaudo_semanal_base":     RECAUDO_SEMANAL_BASE,
        "creditos_activos":         len(loanbooks),
        "creditos_minimos":         creditos_minimos,
        "sobre_el_piso":            len(loanbooks) - creditos_minimos,
        "autosostenible":           len(loanbooks) >= creditos_minimos,
        "total_cartera":            sum(lb.get("saldo_pendiente", 0) or 0 for lb in loanbooks),
        "ticket_promedio":          TICKET_PROMEDIO,
        "cuotas_iniciales_pendientes": {
            "total":   sum(lb.get("cuota_inicial_pendiente", 0) or 0 for lb in cuotas_iniciales_lbs),
            "detalle": [
                {
                    "codigo":  lb.get("codigo"),
                    "cliente": lb.get("cliente_nombre"),
                    "monto":   lb.get("cuota_inicial_pendiente", 0),
                }
                for lb in cuotas_iniciales_lbs
            ],
        },
    }


# ── GET /cfo/plan-deudas ───────────────────────────────────────────────────────
@router.get("/plan-deudas")
async def get_plan_deudas(current_user=Depends(get_current_user)):
    cfg = await db.cfo_financiero_config.find_one({}, {"_id": 0}) or {}
    gastos = cfg.get("gastos_fijos_semanales", 0)

    if gastos == 0:
        return {
            "error": "Configura los gastos_fijos_semanales primero",
            "semanas": [],
        }

    reserva   = gastos * cfg.get("reserva_minima_semanas", 2)
    margen    = RECAUDO_SEMANAL_BASE - gastos - reserva

    deudas_raw = await db.cfo_deudas.find(
        {"tipo": "no_productiva", "estado": {"$ne": "pagada"}}, {"_id": 0}
    ).sort("tasa_mensual", -1).to_list(100)

    if not deudas_raw:
        return {
            "mensaje":        "No hay deudas no productivas. ¡Excelente!",
            "semanas":        [],
            "margen_semanal": margen,
            "gastos_fijos":   gastos,
            "reserva":        reserva,
        }

    if margen <= 0:
        return {
            "error":   f"Margen disponible: {_fmt_cop(margen)}. Los gastos superan el recaudo.",
            "semanas": [],
            "margen_semanal": margen,
        }

    today = date.today()
    saldos = {d["id"]: d.get("saldo_pendiente", d.get("monto_total", 0)) for d in deudas_raw}
    semanas = []
    w = 0

    while sum(saldos.values()) > 0 and w < 52:
        wed = _next_wednesday(today + timedelta(weeks=w))
        disponible = margen
        pagos = []

        for d in deudas_raw:
            did = d["id"]
            if saldos.get(did, 0) <= 0:
                continue
            pago = min(saldos[did], disponible)
            if pago > 0:
                pagos.append({"acreedor": d["acreedor"], "monto": round(pago), "id": did})
                saldos[did] -= pago
                disponible  -= pago
            if disponible <= 0:
                break

        semanas.append({
            "semana":             w + 1,
            "miercoles":          wed.isoformat(),
            "recaudo":            RECAUDO_SEMANAL_BASE,
            "gastos_fijos":       gastos,
            "reserva":            reserva,
            "disponible_deuda":   round(margen),
            "pagos":              pagos,
            "deuda_np_restante":  round(sum(saldos.values())),
        })
        w += 1

    total_np = sum(d.get("saldo_pendiente", d.get("monto_total", 0)) for d in deudas_raw)
    return {
        "semanas":            semanas,
        "total_np":           total_np,
        "margen_semanal":     round(margen),
        "gastos_fijos":       gastos,
        "reserva":            reserva,
        "recaudo_base":       RECAUDO_SEMANAL_BASE,
        "semanas_liberacion": len(semanas),
        "fecha_liberacion":   (_next_wednesday(today + timedelta(weeks=len(semanas)))).isoformat() if semanas else None,
        "metodo":             "AVALANCHA (mayor tasa primero)",
    }


# ── GET /cfo/cuotas-iniciales ──────────────────────────────────────────────────
@router.get("/cuotas-iniciales")
async def get_cuotas_iniciales(current_user=Depends(get_current_user)):
    lbs = await db.loanbook.find(
        {"cuota_inicial_pendiente": {"$gt": 0}},
        {"_id": 0, "codigo": 1, "cliente_nombre": 1,
         "cuota_inicial_total": 1, "cuota_inicial_pendiente": 1,
         "cuota_inicial_pagada": 1, "factura_alegra_id": 1}
    ).to_list(50)

    total = sum(lb.get("cuota_inicial_pendiente", 0) for lb in lbs)
    return {
        "total_pendiente": total,
        "detalle": [
            {
                "codigo":    lb["codigo"],
                "cliente":   lb["cliente_nombre"],
                "total":     lb.get("cuota_inicial_total", 0),
                "pagado":    (lb.get("cuota_inicial_total", 0) or 0) - (lb.get("cuota_inicial_pendiente", 0) or 0),
                "pendiente": lb.get("cuota_inicial_pendiente", 0),
                "factura":   f"FV-{lb.get('factura_alegra_id', '')}",
            }
            for lb in lbs
        ],
    }


# ── GET /cfo/indicadores ───────────────────────────────────────────────────────
@router.get("/indicadores")
async def get_indicadores(current_user=Depends(get_current_user)):
    cfg = await db.cfo_financiero_config.find_one({}, {"_id": 0}) or {}
    gastos = cfg.get("gastos_fijos_semanales", 0)

    activos = await db.loanbook.count_documents({"estado": "activo"})
    saldo_total = 0
    lbs = await db.loanbook.find(
        {"estado": "activo"}, {"_id": 0, "saldo_pendiente": 1, "cuota_valor": 1}
    ).to_list(100)
    saldo_total = sum(lb.get("saldo_pendiente", 0) or 0 for lb in lbs)

    deuda_np = await db.cfo_deudas.aggregate([
        {"$match": {"tipo": "no_productiva", "estado": {"$ne": "pagada"}}},
        {"$group": {"_id": None, "total": {"$sum": "$saldo_pendiente"}}}
    ]).to_list(1)
    total_np = deuda_np[0]["total"] if deuda_np else 0

    deuda_p = await db.cfo_deudas.aggregate([
        {"$match": {"tipo": "productiva", "estado": {"$ne": "pagada"}}},
        {"$group": {"_id": None, "total": {"$sum": "$saldo_pendiente"}}}
    ]).to_list(1)
    total_p = deuda_p[0]["total"] if deuda_p else 0

    creditos_min = int(-(-gastos // TICKET_PROMEDIO)) if gastos > 0 else 0
    margen = RECAUDO_SEMANAL_BASE - gastos - (gastos * 2) if gastos > 0 else 0
    pct_gastos = ((gastos / RECAUDO_SEMANAL_BASE) * 100) if RECAUDO_SEMANAL_BASE > 0 else 0

    return {
        "recaudo_semanal_base":  RECAUDO_SEMANAL_BASE,
        "creditos_activos":      activos,
        "creditos_minimos":      creditos_min,
        "sobre_el_piso":         activos - creditos_min,
        "autosostenible":        activos >= creditos_min,
        "saldo_cartera":         saldo_total,
        "deuda_no_productiva":   total_np,
        "deuda_productiva":      total_p,
        "margen_semanal":        round(margen),
        "pct_gastos_vs_recaudo": round(pct_gastos, 1),
        "gastos_fijos_config":   gastos,
        "configurado":           gastos > 0,
    }


# ── GET /cfo/reporte-lunes ─────────────────────────────────────────────────────
@router.get("/reporte-lunes")
async def get_reporte_lunes(current_user=Depends(get_current_user)):
    """Reporte CFO del lunes — usa datos reales de MongoDB."""
    today     = date.today()
    this_wed  = _next_wednesday(today)
    next_wed  = this_wed + timedelta(weeks=1)

    cfg    = await db.cfo_financiero_config.find_one({}, {"_id": 0}) or {}
    gastos = cfg.get("gastos_fijos_semanales", 0)

    lbs = await db.loanbook.find(
        {"estado": "activo", "cuota_valor": {"$gt": 0}}, {"_id": 0}
    ).to_list(100)

    # Cuotas esta semana
    cuotas_semana = []
    for lb in lbs:
        cv = lb.get("cuota_valor", 0) or 0
        for c in lb.get("cuotas", []):
            if c.get("estado") == "pendiente":
                fv = c.get("fecha_vencimiento", "")
                if fv:
                    try:
                        d = date.fromisoformat(fv[:10])
                        if today <= d < next_wed:
                            cuotas_semana.append({"valor": cv, "cliente": lb.get("cliente_nombre", ""), "fecha": fv[:10]})
                    except Exception:
                        pass

    recaudo_sem = sum(c["valor"] for c in cuotas_semana)

    ci_lbs = await db.loanbook.find(
        {"cuota_inicial_pendiente": {"$gt": 0}}, {"_id": 0}
    ).to_list(50)
    ci_pendiente = sum(lb.get("cuota_inicial_pendiente", 0) for lb in ci_lbs)

    # Deudas
    deuda_np_docs = await db.cfo_deudas.find(
        {"tipo": "no_productiva", "estado": {"$ne": "pagada"}}, {"_id": 0}
    ).to_list(50)
    total_np = sum(d.get("saldo_pendiente", 0) for d in deuda_np_docs)

    # Plan deudas — pago programado esta semana
    pago_deuda_sem = 0
    if gastos > 0:
        reserva = gastos * cfg.get("reserva_minima_semanas", 2)
        margen  = RECAUDO_SEMANAL_BASE - gastos - reserva
        pago_deuda_sem = max(0, min(margen, total_np))

    caja_proyectada = recaudo_sem - gastos - pago_deuda_sem
    activos = len(lbs)
    creditos_min = int(-(-gastos // TICKET_PROMEDIO)) if gastos > 0 else 0

    estado_caja  = "verde"  if caja_proyectada > 0          else "rojo"
    estado_cred  = "verde"  if activos >= creditos_min       else ("amarillo" if activos >= creditos_min * 0.8 else "rojo")
    ci_alerta    = ci_pendiente > 0 and (today - date(2026, 3, 1)).days > 30

    return {
        "fecha":            today.isoformat(),
        "semana_label":     f"Semana del {today.strftime('%d-%b-%Y')}",
        "miercoles":        this_wed.isoformat(),
        "ingresos": {
            "recaudo_cartera":    recaudo_sem,
            "num_cuotas":         len(cuotas_semana),
            "ci_pendiente":       ci_pendiente,
            "ci_detalle":         [{"cliente": lb.get("cliente_nombre"), "monto": lb.get("cuota_inicial_pendiente", 0)} for lb in ci_lbs],
        },
        "egresos": {
            "gastos_fijos":       gastos,
            "pago_deuda_semana":  pago_deuda_sem,
            "total_comprometido": gastos + pago_deuda_sem,
        },
        "caja": {
            "proyectada":  caja_proyectada,
            "estado":      estado_caja,
        },
        "deuda": {
            "no_productiva": total_np,
            "estado":        "verde" if total_np == 0 else ("amarillo" if total_np < 5_000_000 else "rojo"),
        },
        "creditos": {
            "activos":        activos,
            "minimo":         creditos_min,
            "sobre_piso":     activos - creditos_min,
            "estado":         estado_cred,
        },
        "alertas": [
            *([{"tipo": "caja_negativa", "msg": f"⚠️ Caja proyectada negativa: ${caja_proyectada:,.0f}"}] if caja_proyectada < 0 else []),
            *([{"tipo": "bajo_piso", "msg": f"⚠️ Créditos activos ({activos}) por debajo del mínimo ({creditos_min})"}] if activos < creditos_min else []),
            *([{"tipo": "ci_sin_cobrar", "msg": f"⚠️ Cuotas iniciales sin cobrar: ${ci_pendiente:,.0f} (>30 días)"}] if ci_alerta else []),
            *([{"tipo": "deuda_np_alta", "msg": f"⚠️ Deuda no productiva: ${total_np:,.0f}. Prioriza liquidarla."}] if total_np > 0 else []),
        ],
    }
