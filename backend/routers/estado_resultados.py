"""estado_resultados.py — BUILD 12: Estado de Resultados (P&L) RODDOS.

Endpoints:
  GET  /cfo/estado-resultados          — P&L completo (JSON)
  GET  /cfo/estado-resultados/pdf      — Exportar PDF ejecutivo
  GET  /cfo/estado-resultados/excel    — Exportar Excel 4 hojas
"""
import io
import logging
from datetime import datetime, timezone, date, timedelta
from typing import Optional

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side, numbers
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse

from database import db
from dependencies import get_current_user
from alegra_service import AlegraService

logger = logging.getLogger(__name__)
router = APIRouter(prefix="/cfo", tags=["Estado de Resultados"])

CORP_BLUE   = "0F2A5C"
CORP_GOLD   = "C9A84C"
CORP_LIGHT  = "D9E1F2"
TAX_RATE    = 0.33

MESES_ES = ["enero","febrero","marzo","abril","mayo","junio",
            "julio","agosto","septiembre","octubre","noviembre","diciembre"]

# ── Category keyword classifier ───────────────────────────────────────────────
def _classify_bill(proveedor: str, desc: str) -> str:
    txt = f"{proveedor} {desc}".lower()
    if any(k in txt for k in ["nomina","nómina","personal","salario","prestacion","prestación","sueldo"]):
        return "personal"
    if any(k in txt for k in ["arriendo","arrendamiento","canon","local"]):
        return "arrendamiento"
    if any(k in txt for k in ["servicio","energia","energía","agua","gas","acueducto","internet","telefono","teléfono","claro","movistar","une"]):
        return "servicios"
    if any(k in txt for k in ["honorario","contable","asesor","juridico","jurídico","tributar"]):
        return "honorarios"
    if any(k in txt for k in ["retencion","retención","dian"]):
        return "retenciones"
    return "otros"


def _fmt_cop(n: float) -> str:
    return f"${n:,.0f}"


def _period_bounds(periodo: str) -> tuple[str, str]:
    """YYYY-MM → (first_day, last_day) as ISO strings."""
    ano, mes = int(periodo[:4]), int(periodo[5:7])
    first = date(ano, mes, 1)
    if mes == 12:
        last = date(ano + 1, 1, 1) - timedelta(days=1)
    else:
        last = date(ano, mes + 1, 1) - timedelta(days=1)
    return first.isoformat(), last.isoformat()


def _prev_period(periodo: str) -> str:
    ano, mes = int(periodo[:4]), int(periodo[5:7])
    mes -= 1
    if mes == 0:
        mes = 12
        ano -= 1
    return f"{ano}-{mes:02d}"


# ── Core P&L builder ──────────────────────────────────────────────────────────
async def _build_pl(periodo: str, user):
    """Build the full P&L dict for a given YYYY-MM period."""
    alegra = AlegraService(db)
    first_day, last_day = _period_bounds(periodo)
    ano, mes = int(periodo[:4]), int(periodo[5:7])
    mes_label = f"{MESES_ES[mes-1].capitalize()} {ano}"

    # ── 1. Ingresos: Alegra invoices ──────────────────────────────────────────
    invoices_raw: list = []
    try:
        invoices_raw = await alegra.request("invoices") or []
        invoices = [
            inv for inv in invoices_raw
            if first_day <= (inv.get("date") or "9999")[:10] <= last_day
            and inv.get("status") not in ("void", "draft")
        ]
    except Exception as e:
        logger.warning("estado_resultados invoices error: %s", e)
        invoices = []

    ventas_motos_n_total = 0
    ventas_motos_n_units = 0
    ventas_motos_u_total = 0
    ventas_motos_u_units = 0
    ventas_repuestos     = 0
    ingresos_financiacion= 0
    otros_ingresos       = 0
    detalle_ingresos     = []

    for inv in invoices:
        total = float(inv.get("total") or 0)
        cliente = (inv.get("client") or {}).get("name") or inv.get("client_name") or "—"
        inv_date = (inv.get("date") or "")[:10]
        items = inv.get("items") or []
        desc_items = " ".join(str(it.get("name") or it.get("description") or "") for it in items).lower()

        # Classify by invoice items/description
        if any(k in desc_items for k in ["moto","honda","bajaj","yamaha","tvs","kawasaki","auteco","suzuki","akt"]):
            # Try to distinguish new vs used
            is_used = any(k in desc_items for k in ["usado","usada","seminuevo","seminueva","second"])
            if is_used:
                ventas_motos_u_total += total
                ventas_motos_u_units += 1
            else:
                ventas_motos_n_total += total
                ventas_motos_n_units += 1
        elif any(k in desc_items for k in ["repuesto","accesorio","casco","llanta","aceite","pieza"]):
            ventas_repuestos += total
        elif any(k in desc_items for k in ["interes","interés","financiaci","cuota"]):
            ingresos_financiacion += total
        else:
            otros_ingresos += total

        detalle_ingresos.append({
            "factura":  inv.get("number") or inv.get("numberTemplate", {}).get("number") or f"#{inv.get('id','')}",
            "fecha":    inv_date,
            "cliente":  cliente,
            "total":    total,
            "concepto": desc_items[:60] or "—",
        })

    total_ingresos = ventas_motos_n_total + ventas_motos_u_total + ventas_repuestos + ingresos_financiacion + otros_ingresos

    # ── 2. Costo de ventas: motos vendidas en el período ─────────────────────
    motos_vendidas = await db.inventario_motos.find(
        {
            "estado": {"$in": ["Vendida", "Entregada"]},
            "$or": [
                {"fecha_venta":    {"$gte": first_day, "$lte": last_day}},
                {"updated_at":     {"$gte": first_day, "$lte": last_day + "T23:59:59"}},
            ],
        },
        {"_id": 0},
    ).to_list(200)

    costo_motos_n = 0.0
    costo_motos_n_units = 0
    costo_motos_u = 0.0
    costo_motos_u_units = 0
    sin_costo = 0
    for m in motos_vendidas:
        costo = float(m.get("precio_compra") or 0) + float(m.get("iva_compra") or 0) + float(m.get("ipoc_compra") or 0)
        is_used = "usad" in (m.get("referencia") or "").lower() or "usad" in (m.get("descripcion") or "").lower()
        if costo == 0:
            sin_costo += 1
        if is_used:
            costo_motos_u += costo
            costo_motos_u_units += 1
        else:
            costo_motos_n += costo
            costo_motos_n_units += 1

    total_cogs = costo_motos_n + costo_motos_u
    advertencia_cogs = (
        f"⚠️ {sin_costo} de {len(motos_vendidas)} motos no tienen costo de compra registrado. "
        "Usa 'Carga inicial de costos' en el módulo Motos."
    ) if sin_costo > 0 else None

    utilidad_bruta = total_ingresos - total_cogs
    margen_bruto_pct = (utilidad_bruta / total_ingresos * 100) if total_ingresos > 0 else 0

    # ── 3. Gastos operacionales: Alegra bills ─────────────────────────────────
    try:
        bills_raw = await alegra.request("bills") or []
        bills = [
            b for b in bills_raw
            if first_day <= (b.get("date") or "9999")[:10] <= last_day
        ]
    except Exception as e:
        logger.warning("estado_resultados bills error: %s", e)
        bills = []

    gastos_por_cat: dict[str, float] = {
        "personal": 0, "arrendamiento": 0, "servicios": 0,
        "honorarios": 0, "retenciones": 0, "otros": 0,
    }
    detalle_gastos = []
    for b in bills:
        monto   = float(b.get("total") or 0)
        if monto > 15_000_000:          # skip moto inventory purchases
            continue
        proveedor = (b.get("vendor") or b.get("provider") or {}).get("name") or "—"
        desc      = str(b.get("description") or b.get("observations") or "")
        cat       = _classify_bill(proveedor, desc)
        gastos_por_cat[cat] += monto
        detalle_gastos.append({
            "fecha":    (b.get("date") or "")[:10],
            "proveedor": proveedor,
            "concepto": desc[:60] or cat,
            "cuenta":   cat,
            "total":    monto,
        })

    total_gastos_op = sum(gastos_por_cat.values())
    modo_parcial_gastos = total_gastos_op == 0
    advertencia_gastos  = (
        f"⚠️ Gastos de {mes_label} pendientes de registrar en Alegra. "
        "El P&L se completará automáticamente cuando se carguen."
    ) if modo_parcial_gastos else None

    utilidad_op = utilidad_bruta - total_gastos_op

    # ── 4. Gastos no operacionales: intereses de deuda ───────────────────────
    deudas = await db.cfo_deudas.find(
        {"estado": {"$ne": "pagada"}}, {"_id": 0}
    ).to_list(50)
    intereses = sum(
        float(d.get("saldo_pendiente") or 0) * float(d.get("tasa_mensual") or 0) / 100
        for d in deudas
    )
    total_no_op = intereses

    utilidad_ai = utilidad_op - total_no_op
    provision_imp = max(0, utilidad_ai * TAX_RATE)
    utilidad_neta = utilidad_ai - provision_imp

    # ── 5. Comparativo mes anterior — shallow (no recursion) ──────────────────
    comparativo = None
    try:
        prev_per    = _prev_period(periodo)
        p_first, p_last = _period_bounds(prev_per)
        prev_invoices = [
            inv for inv in invoices_raw  # reuse same fetch
            if p_first <= (inv.get("date") or "9999")[:10] <= p_last
            and inv.get("status") not in ("void", "draft")
        ] if invoices_raw else []
        ing_prev = sum(float(inv.get("total") or 0) for inv in prev_invoices)
        comparativo = {
            "periodo_anterior": prev_per,
            "ingresos_anterior": ing_prev,
            "variacion_ingresos_pct": round((total_ingresos - ing_prev) / max(1, ing_prev) * 100, 1) if ing_prev else None,
        }
    except Exception:
        comparativo = None

    modo = "parcial" if (modo_parcial_gastos or not invoices) else "completo"

    return {
        "periodo":      periodo,
        "mes_label":    mes_label,
        "modo":         modo,
        "generado_en":  datetime.now(timezone.utc).isoformat(),
        "ingresos": {
            "ventas_motos_nuevas":    {"total": ventas_motos_n_total, "unidades": ventas_motos_n_units},
            "ventas_motos_usadas":    {"total": ventas_motos_u_total, "unidades": ventas_motos_u_units},
            "ventas_repuestos":       {"total": ventas_repuestos},
            "ingresos_financiacion":  {"total": ingresos_financiacion},
            "otros_ingresos":         {"total": otros_ingresos},
            "total":                  total_ingresos,
            "detalle":                detalle_ingresos,
        },
        "costo_ventas": {
            "costo_motos_nuevas":  {"total": costo_motos_n, "unidades": costo_motos_n_units},
            "costo_motos_usadas":  {"total": costo_motos_u, "unidades": costo_motos_u_units},
            "total":               total_cogs,
            "motos_sin_costo":     sin_costo,
            "advertencia":         advertencia_cogs,
        },
        "utilidad_bruta":     utilidad_bruta,
        "margen_bruto_pct":   round(margen_bruto_pct, 1),
        "gastos_operacionales": {
            **gastos_por_cat,
            "total":            total_gastos_op,
            "modo_parcial":     modo_parcial_gastos,
            "advertencia":      advertencia_gastos,
            "detalle":          detalle_gastos,
        },
        "utilidad_operacional":     utilidad_op,
        "gastos_no_operacionales": {
            "intereses_deuda":  intereses,
            "otros_financieros": 0,
            "total":            total_no_op,
        },
        "utilidad_antes_impuestos": utilidad_ai,
        "provision_impuestos":      provision_imp,
        "utilidad_neta":            utilidad_neta,
        "alerta_margen_critico":    (margen_bruto_pct < 15 and total_ingresos > 0),
        "comparativo":              comparativo,
    }


# ── GET /cfo/estado-resultados ────────────────────────────────────────────────
@router.get("/estado-resultados")
async def get_estado_resultados(
    periodo: Optional[str] = None,
    current_user=Depends(get_current_user),
):
    """Retorna el P&L completo del período (YYYY-MM). Default: mes actual."""
    if not periodo:
        hoy = date.today()
        periodo = f"{hoy.year}-{hoy.month:02d}"
    try:
        pl = await _build_pl(periodo, current_user)
        return pl
    except Exception as e:
        logger.error("estado_resultados error: %s", e)
        raise HTTPException(status_code=500, detail=str(e))


# ── GET /cfo/estado-resultados/excel ─────────────────────────────────────────
@router.get("/estado-resultados/excel")
async def export_estado_resultados_excel(
    periodo: Optional[str] = None,
    current_user=Depends(get_current_user),
):
    """Exporta el P&L en Excel con 4 hojas."""
    if not periodo:
        hoy = date.today()
        periodo = f"{hoy.year}-{hoy.month:02d}"

    pl = await _build_pl(periodo, current_user)
    wb = _build_excel(pl)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"RODDOS_PL_{periodo}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


# ── GET /cfo/estado-resultados/pdf ───────────────────────────────────────────
@router.get("/estado-resultados/pdf")
async def export_estado_resultados_pdf(
    periodo: Optional[str] = None,
    current_user=Depends(get_current_user),
):
    """Exporta el P&L en PDF ejecutivo con análisis automático."""
    if not periodo:
        hoy = date.today()
        periodo = f"{hoy.year}-{hoy.month:02d}"

    pl = await _build_pl(periodo, current_user)
    buf = _build_pdf(pl)
    fname = f"RODDOS_PL_{periodo}.pdf"
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


# ── Excel builder ─────────────────────────────────────────────────────────────
def _build_excel(pl: dict) -> openpyxl.Workbook:
    wb = openpyxl.Workbook()

    thin = Side(style="thin", color="CCCCCC")
    brd  = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_fill  = PatternFill("solid", fgColor=CORP_BLUE)
    hdr_font  = Font(bold=True, color="FFFFFF", size=11)
    sub_fill  = PatternFill("solid", fgColor=CORP_LIGHT)
    tot_font  = Font(bold=True, size=11)
    gold_fill = PatternFill("solid", fgColor="FFF8E7")
    red_font  = Font(bold=True, color="C0392B", size=11)
    grn_font  = Font(bold=True, color="1E8449", size=11)

    def _hdr(ws, row, cols):
        for col, val in enumerate(cols, 1):
            c = ws.cell(row=row, column=col, value=val)
            c.fill = hdr_fill; c.font = hdr_font; c.border = brd
            c.alignment = Alignment(horizontal="center", vertical="center")

    def _row(ws, r, cols, bold=False, fill=None, right_cols=None):
        right_cols = right_cols or []
        for col, val in enumerate(cols, 1):
            c = ws.cell(row=r, column=col, value=val)
            c.border = brd
            if bold: c.font = tot_font
            if fill: c.fill = fill
            if col in right_cols: c.alignment = Alignment(horizontal="right")

    # ── Hoja 1: P&L ──────────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "P&L"

    ws1.merge_cells("A1:D1")
    t = ws1["A1"]
    t.value = f"ESTADO DE RESULTADOS — {pl['mes_label'].upper()}"
    t.font = Font(bold=True, color=CORP_BLUE, size=14)
    t.alignment = Alignment(horizontal="center")

    ws1.merge_cells("A2:D2")
    ws1["A2"].value = "RODDOS S.A.S."
    ws1["A2"].font = Font(bold=True, size=11, color="555555")
    ws1["A2"].alignment = Alignment(horizontal="center")

    ws1.merge_cells("A3:D3")
    ws1["A3"].value = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}  |  Modo: {pl['modo'].upper()}"
    ws1["A3"].font = Font(size=9, color="888888")
    ws1["A3"].alignment = Alignment(horizontal="center")

    _hdr(ws1, 5, ["CONCEPTO", "MES ACTUAL", "MES ANTERIOR", "VARIACIÓN %"])

    rows_pl = [
        ("INGRESOS OPERACIONALES", "", "", "", True, hdr_fill),
        ("  Ventas motos nuevas",  pl["ingresos"]["ventas_motos_nuevas"]["total"],
         pl.get("comparativo",{}) and None, None, False, None),
        ("  Ventas motos usadas",  pl["ingresos"]["ventas_motos_usadas"]["total"], None, None, False, None),
        ("  Ventas repuestos",     pl["ingresos"]["ventas_repuestos"]["total"], None, None, False, None),
        ("  Ingresos financiación",pl["ingresos"]["ingresos_financiacion"]["total"], None, None, False, None),
        ("  Otros ingresos",       pl["ingresos"]["otros_ingresos"]["total"], None, None, False, None),
        ("TOTAL INGRESOS",         pl["ingresos"]["total"],
         pl.get("comparativo",{}) and pl["comparativo"]["ingresos_anterior"] if pl.get("comparativo") else 0,
         pl["comparativo"]["variacion_ingresos_pct"] if pl.get("comparativo") else None,
         True, sub_fill),
        ("", "", "", "", False, None),
        ("COSTO DE VENTAS", "", "", "", True, hdr_fill),
        ("  Costo motos nuevas",   pl["costo_ventas"]["costo_motos_nuevas"]["total"], None, None, False, None),
        ("  Costo motos usadas",   pl["costo_ventas"]["costo_motos_usadas"]["total"], None, None, False, None),
        ("TOTAL COSTO DE VENTAS",  pl["costo_ventas"]["total"], None, None, True, sub_fill),
        ("UTILIDAD BRUTA",         pl["utilidad_bruta"], None,
         f"{pl['margen_bruto_pct']}%", True, gold_fill),
        ("", "", "", "", False, None),
        ("GASTOS OPERACIONALES", "", "", "", True, hdr_fill),
        ("  Personal / nómina",    pl["gastos_operacionales"]["personal"], None, None, False, None),
        ("  Arrendamiento",        pl["gastos_operacionales"]["arrendamiento"], None, None, False, None),
        ("  Servicios públicos",   pl["gastos_operacionales"]["servicios"], None, None, False, None),
        ("  Honorarios",           pl["gastos_operacionales"]["honorarios"], None, None, False, None),
        ("  Retenciones",          pl["gastos_operacionales"]["retenciones"], None, None, False, None),
        ("  Otros",                pl["gastos_operacionales"]["otros"], None, None, False, None),
        ("TOTAL GASTOS OPER.",     pl["gastos_operacionales"]["total"], None, None, True, sub_fill),
        ("UTILIDAD OPERACIONAL",   pl["utilidad_operacional"], None, None, True, gold_fill),
        ("", "", "", "", False, None),
        ("GASTOS NO OPERACIONALES","", "", "", True, hdr_fill),
        ("  Intereses de deuda",   pl["gastos_no_operacionales"]["intereses_deuda"], None, None, False, None),
        ("  Otros financieros",    pl["gastos_no_operacionales"]["otros_financieros"], None, None, False, None),
        ("TOTAL NO OPERACIONAL",   pl["gastos_no_operacionales"]["total"], None, None, True, sub_fill),
        ("", "", "", "", False, None),
        ("UTILIDAD ANTES IMPUESTO",pl["utilidad_antes_impuestos"], None, None, True, gold_fill),
        ("  Provisión impuestos (33%)", pl["provision_impuestos"], None, None, False, None),
        ("UTILIDAD NETA",          pl["utilidad_neta"], None, None, True, gold_fill),
    ]

    for i, row_data in enumerate(rows_pl, start=6):
        concepto, val, prev_val, var, bold, fill = row_data
        ws1.cell(i, 1, concepto).border = brd
        for col, v in [(2, val), (3, prev_val), (4, var)]:
            c = ws1.cell(i, col, v if v != "" else None)
            c.border = brd
            if isinstance(v, (int, float)) and v is not None:
                c.number_format = '#,##0'
                c.alignment = Alignment(horizontal="right")
        if bold:
            for col in range(1, 5):
                ws1.cell(i, col).font = tot_font
        if fill:
            for col in range(1, 5):
                ws1.cell(i, col).fill = fill

    # Final row — utilidad neta color
    net_row = 6 + len(rows_pl) - 1
    net_font = grn_font if pl["utilidad_neta"] >= 0 else red_font
    ws1.cell(net_row, 1).font = net_font
    ws1.cell(net_row, 2).font = net_font

    for i, w in enumerate([38, 18, 18, 14], 1):
        ws1.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # ── Hoja 2: Detalle Ingresos ──────────────────────────────────────────────
    ws2 = wb.create_sheet("Detalle Ingresos")
    _hdr(ws2, 1, ["Factura", "Fecha", "Cliente", "Total", "Concepto"])
    for r, d in enumerate(pl["ingresos"]["detalle"], 2):
        ws2.cell(r, 1, d["factura"])
        ws2.cell(r, 2, d["fecha"])
        ws2.cell(r, 3, d["cliente"])
        c = ws2.cell(r, 4, d["total"])
        c.number_format = '#,##0'
        c.alignment = Alignment(horizontal="right")
        ws2.cell(r, 5, d["concepto"])
    for i, w in enumerate([14, 12, 30, 15, 40], 1):
        ws2.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # ── Hoja 3: Detalle Gastos ────────────────────────────────────────────────
    ws3 = wb.create_sheet("Detalle Gastos")
    _hdr(ws3, 1, ["Fecha", "Proveedor", "Concepto", "Cuenta", "Total"])
    for r, d in enumerate(pl["gastos_operacionales"]["detalle"], 2):
        ws3.cell(r, 1, d["fecha"])
        ws3.cell(r, 2, d["proveedor"])
        ws3.cell(r, 3, d["concepto"])
        ws3.cell(r, 4, d["cuenta"])
        c = ws3.cell(r, 5, d["total"])
        c.number_format = '#,##0'
        c.alignment = Alignment(horizontal="right")
    for i, w in enumerate([12, 28, 40, 16, 15], 1):
        ws3.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # ── Hoja 4: Cartera ───────────────────────────────────────────────────────
    ws4 = wb.create_sheet("Cartera")
    _hdr(ws4, 1, ["Código", "Cliente", "Plan", "Cuota/sem", "Cuotas Pag.", "Cuotas Pend.", "Saldo", "Próxima cuota"])
    lbs_placeholder = [("Cargando…", "", "", 0, 0, 0, 0, "")]
    for i, w in enumerate([14, 30, 8, 12, 12, 12, 16, 14], 1):
        ws4.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    return wb


# ── PDF builder ───────────────────────────────────────────────────────────────
def _build_pdf(pl: dict) -> io.BytesIO:
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.units import cm
    from reportlab.lib.colors import HexColor, white, black
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=letter,
                            leftMargin=2*cm, rightMargin=2*cm,
                            topMargin=2*cm, bottomMargin=2*cm)

    BLUE  = HexColor(f"#{CORP_BLUE}")
    GOLD  = HexColor(f"#{CORP_GOLD}")
    LIGHT = HexColor("#EEF2FF")
    RED   = HexColor("#C0392B")
    GREEN = HexColor("#1E8449")

    styles = getSampleStyleSheet()
    title_style  = ParagraphStyle("Title",  fontName="Helvetica-Bold", fontSize=16, textColor=BLUE,  alignment=TA_CENTER, spaceAfter=4)
    sub_style    = ParagraphStyle("Sub",    fontName="Helvetica",      fontSize=10, textColor=black, alignment=TA_CENTER, spaceAfter=2)
    note_style   = ParagraphStyle("Note",   fontName="Helvetica",      fontSize=8,  textColor=HexColor("#888888"), alignment=TA_CENTER, spaceAfter=12)
    section_style= ParagraphStyle("Sec",    fontName="Helvetica-Bold", fontSize=11, textColor=BLUE,  spaceAfter=6)
    body_style   = ParagraphStyle("Body",   fontName="Helvetica",      fontSize=9,  textColor=black, spaceAfter=4)

    def _fmt(n):
        if n is None: return "—"
        neg = n < 0
        s = f"${abs(round(n)):,}"
        return f"({s})" if neg else s

    elems = []

    # Header
    elems.append(Paragraph("ESTADO DE RESULTADOS", title_style))
    elems.append(Paragraph(f"{pl['mes_label'].upper()}  ·  RODDOS S.A.S.", sub_style))
    elems.append(Paragraph(f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}  ·  Modo: {pl['modo'].upper()}", note_style))

    # P&L table
    def row(concepto, valor, bold=False, bg=None, color=None):
        return [concepto, _fmt(valor) if isinstance(valor, (int, float)) else (valor or "")]

    data = [
        ["CONCEPTO", "MONTO"],
        ["INGRESOS OPERACIONALES", ""],
        [f"  Ventas motos nuevas ({pl['ingresos']['ventas_motos_nuevas']['unidades']} un.)",
         _fmt(pl["ingresos"]["ventas_motos_nuevas"]["total"])],
        ["  Ventas motos usadas", _fmt(pl["ingresos"]["ventas_motos_usadas"]["total"])],
        ["  Ventas repuestos",    _fmt(pl["ingresos"]["ventas_repuestos"]["total"])],
        ["  Ingresos financiación",_fmt(pl["ingresos"]["ingresos_financiacion"]["total"])],
        ["  Otros ingresos",      _fmt(pl["ingresos"]["otros_ingresos"]["total"])],
        ["TOTAL INGRESOS",        _fmt(pl["ingresos"]["total"])],
        ["", ""],
        ["COSTO DE VENTAS", ""],
        ["  Costo motos nuevas",  _fmt(pl["costo_ventas"]["costo_motos_nuevas"]["total"])],
        ["  Costo motos usadas",  _fmt(pl["costo_ventas"]["costo_motos_usadas"]["total"])],
        ["TOTAL COSTO DE VENTAS", _fmt(pl["costo_ventas"]["total"])],
        [f"UTILIDAD BRUTA  ({pl['margen_bruto_pct']}%)", _fmt(pl["utilidad_bruta"])],
        ["", ""],
        ["GASTOS OPERACIONALES", ""],
        ["  Personal / nómina",   _fmt(pl["gastos_operacionales"]["personal"])],
        ["  Arrendamiento",       _fmt(pl["gastos_operacionales"]["arrendamiento"])],
        ["  Servicios públicos",  _fmt(pl["gastos_operacionales"]["servicios"])],
        ["  Honorarios",          _fmt(pl["gastos_operacionales"]["honorarios"])],
        ["  Otros",               _fmt(pl["gastos_operacionales"]["otros"])],
        ["TOTAL GASTOS OPER.",    _fmt(pl["gastos_operacionales"]["total"])],
        ["UTILIDAD OPERACIONAL",  _fmt(pl["utilidad_operacional"])],
        ["", ""],
        ["GASTOS NO OPERACIONALES",""],
        ["  Intereses de deuda",  _fmt(pl["gastos_no_operacionales"]["intereses_deuda"])],
        ["TOTAL NO OPERACIONAL",  _fmt(pl["gastos_no_operacionales"]["total"])],
        ["", ""],
        ["UTILIDAD ANTES IMPUESTO",_fmt(pl["utilidad_antes_impuestos"])],
        ["  Provisión impuestos (33%)", _fmt(pl["provision_impuestos"])],
        ["UTILIDAD NETA",         _fmt(pl["utilidad_neta"])],
    ]

    # Indices of "header/section" rows and "total" rows
    section_rows = {0, 1, 9, 15, 24}
    total_rows   = {7, 12, 13, 21, 22, 26, 28, 30}
    net_row_idx  = 30

    t = Table(data, colWidths=[12*cm, 5*cm])
    style_cmds = [
        ("FONTNAME",    (0,0), (-1,0), "Helvetica-Bold"),
        ("FONTSIZE",    (0,0), (-1,-1), 9),
        ("BACKGROUND",  (0,0), (-1,0), BLUE),
        ("TEXTCOLOR",   (0,0), (-1,0), white),
        ("ALIGN",       (1,0), (1,-1), "RIGHT"),
        ("ROWBACKGROUNDS", (0,1), (-1,-1), [white, HexColor("#F8F9FA")]),
        ("GRID",        (0,0), (-1,-1), 0.3, HexColor("#CCCCCC")),
        ("TOPPADDING",  (0,0), (-1,-1), 3),
        ("BOTTOMPADDING",(0,0), (-1,-1), 3),
    ]
    for r in section_rows:
        style_cmds += [
            ("BACKGROUND",  (0,r), (-1,r), BLUE),
            ("TEXTCOLOR",   (0,r), (-1,r), white),
            ("FONTNAME",    (0,r), (-1,r), "Helvetica-Bold"),
        ]
    for r in total_rows:
        style_cmds += [
            ("BACKGROUND",  (0,r), (-1,r), LIGHT),
            ("FONTNAME",    (0,r), (-1,r), "Helvetica-Bold"),
        ]
    # Net income row color
    net_color = GREEN if pl["utilidad_neta"] >= 0 else RED
    style_cmds += [
        ("BACKGROUND",  (0, net_row_idx), (-1, net_row_idx), net_color),
        ("TEXTCOLOR",   (0, net_row_idx), (-1, net_row_idx), white),
        ("FONTNAME",    (0, net_row_idx), (-1, net_row_idx), "Helvetica-Bold"),
        ("FONTSIZE",    (0, net_row_idx), (-1, net_row_idx), 11),
    ]

    t.setStyle(TableStyle(style_cmds))
    elems.append(t)
    elems.append(Spacer(1, 0.5*cm))

    # Warnings
    for adv in [pl["costo_ventas"].get("advertencia"), pl["gastos_operacionales"].get("advertencia")]:
        if adv:
            elems.append(Paragraph(adv, body_style))

    # Auto-analysis
    elems.append(Spacer(1, 0.3*cm))
    elems.append(Paragraph("Análisis del período", section_style))

    ing   = pl["ingresos"]["total"]
    margen= pl["margen_bruto_pct"]
    neta  = pl["utilidad_neta"]
    modo  = pl["modo"]

    if ing == 0:
        analysis = "No se registraron ingresos en Alegra para este período."
    else:
        analysis = (
            f"Los ingresos operacionales del período fueron {_fmt(ing)}. "
            f"El margen bruto {'es' if margen > 0 else 'es negativo en'} {abs(margen):.1f}%, "
            f"{'lo que indica una operación eficiente en costos.' if margen >= 20 else 'por debajo del mínimo recomendado del 20%.' if margen < 15 else 'dentro del rango aceptable.'} "
            f"La utilidad neta {'es positiva' if neta >= 0 else 'es negativa'}: {_fmt(neta)}. "
        )
        if modo == "parcial":
            analysis += "Nota: Los gastos operativos no están en Alegra aún — el P&L está en modo parcial."
        comp = pl.get("comparativo")
        if comp and comp.get("variacion_ingresos_pct") is not None:
            var = comp["variacion_ingresos_pct"]
            analysis += f" Variación vs mes anterior: {'+'if var>=0 else ''}{var:.1f}% en ingresos."

    elems.append(Paragraph(analysis, body_style))

    # Footer note
    elems.append(Spacer(1, 0.5*cm))
    elems.append(Paragraph(
        f"Generado por Agente CFO RODDOS  ·  {datetime.now().strftime('%d/%m/%Y %H:%M')}",
        note_style,
    ))

    doc.build(elems)
    buf.seek(0)
    return buf
