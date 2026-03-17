"""gastos.py — BUILD 16: Carga Masiva de Gastos vía Excel.

Endpoints:
  GET  /gastos/plantilla           — Descarga plantilla Excel (12 columnas)
  POST /gastos/cargar              — Parse Excel → validación + preview con retenciones
  POST /gastos/procesar            — Procesa filas en Alegra (journal-entries o bills)
  GET  /gastos/jobs/{job_id}       — Consulta estado de job asíncrono
  GET  /gastos/reporte-errores/{job_id} — Descarga Excel con errores
"""
import io
import uuid
import logging
import unicodedata
from datetime import datetime, timezone, date, timedelta
from typing import Optional, List
from copy import deepcopy

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, BackgroundTasks
from fastapi.responses import StreamingResponse
from pydantic import BaseModel

from alegra_service import AlegraService
from database import db
from dependencies import get_current_user

router = APIRouter(prefix="/gastos", tags=["gastos-masivos"])
logger = logging.getLogger(__name__)

# ── Constantes contables (Alegra IDs internos de RODDOS SAS) ─────────────────

CUENTAS_GASTO = {
    "arrendamiento": {"id": 5480, "nombre": "Arrendamientos (512010)"},
    "servicios":     {"id": 5483, "nombre": "Asistencia técnica / Servicios"},
    "honorarios_pn": {"id": 5475, "nombre": "Asesoría jurídica / Honorarios PN"},
    "honorarios_pj": {"id": 5476, "nombre": "Asesoría financiera / Honorarios PJ"},
    "compras":       {"id": 5497, "nombre": "Útiles y papelería / Compras misc"},
    "nomina":        {"id": 5462, "nombre": "Sueldos y salarios (510506)"},
    "impuesto":      {"id": 5478, "nombre": "Industria y Comercio (511505)"},
    "otros":         {"id": 5493, "nombre": "Gastos generales (5195)"},  # FIXED: was 5495 Gastos de representación
}

# Plan de cuentas RODDOS — mapeado a IDs reales de Alegra
# Usado para categoria+subcategoria del nuevo template
PLAN_CUENTAS_RODDOS = [
    # PERSONAL
    {"categoria": "Personal", "subcategoria": "Salarios",          "alegra_id": 5462, "cuenta_codigo": "510506", "cuenta_nombre": "Sueldos y salarios", "tipo_retefuente": "nomina"},
    {"categoria": "Personal", "subcategoria": "Honorarios",        "alegra_id": 5475, "cuenta_codigo": "511025", "cuenta_nombre": "Honorarios (asesoría)", "tipo_retefuente": "honorarios_pn"},
    {"categoria": "Personal", "subcategoria": "Honorarios_PJ",     "alegra_id": 5476, "cuenta_codigo": "511030", "cuenta_nombre": "Honorarios PJ", "tipo_retefuente": "honorarios_pj"},
    {"categoria": "Personal", "subcategoria": "Seguridad_Social",  "alegra_id": 5472, "cuenta_codigo": "510570", "cuenta_nombre": "Aportes seguridad social", "tipo_retefuente": "nomina"},
    {"categoria": "Personal", "subcategoria": "Dotacion",          "alegra_id": 5470, "cuenta_codigo": "510551", "cuenta_nombre": "Dotación a trabajadores", "tipo_retefuente": "nomina"},
    {"categoria": "Personal", "subcategoria": "Vacaciones",        "alegra_id": 5469, "cuenta_codigo": "510539", "cuenta_nombre": "Vacaciones", "tipo_retefuente": "nomina"},
    {"categoria": "Personal", "subcategoria": "Prima",             "alegra_id": 5468, "cuenta_codigo": "510536", "cuenta_nombre": "Prima de servicios", "tipo_retefuente": "nomina"},
    {"categoria": "Personal", "subcategoria": "Cesantias",         "alegra_id": 5466, "cuenta_codigo": "510530", "cuenta_nombre": "Cesantías", "tipo_retefuente": "nomina"},
    # OPERACIONES
    {"categoria": "Operaciones", "subcategoria": "Arriendo",          "alegra_id": 5480, "cuenta_codigo": "512010", "cuenta_nombre": "Arrendamientos", "tipo_retefuente": "arrendamiento"},
    {"categoria": "Operaciones", "subcategoria": "Servicios_Publicos", "alegra_id": 5485, "cuenta_codigo": "513525", "cuenta_nombre": "Alcantarillado/Acueducto/Servicios públicos", "tipo_retefuente": "servicios"},
    {"categoria": "Operaciones", "subcategoria": "Telefonia",          "alegra_id": 5487, "cuenta_codigo": "513535", "cuenta_nombre": "Teléfono/Internet/Comunicaciones", "tipo_retefuente": "servicios"},
    {"categoria": "Operaciones", "subcategoria": "Mantenimiento",      "alegra_id": 5483, "cuenta_codigo": "513515", "cuenta_nombre": "Asistencia técnica/Mantenimiento", "tipo_retefuente": "servicios"},
    {"categoria": "Operaciones", "subcategoria": "Transporte",         "alegra_id": 5499, "cuenta_codigo": "519545", "cuenta_nombre": "Taxis y buses/Transporte", "tipo_retefuente": "otros"},
    {"categoria": "Operaciones", "subcategoria": "Papeleria",          "alegra_id": 5497, "cuenta_codigo": "519530", "cuenta_nombre": "Útiles, papelería y fotocopia", "tipo_retefuente": "compras"},
    {"categoria": "Operaciones", "subcategoria": "Aseo",               "alegra_id": 5482, "cuenta_codigo": "513505", "cuenta_nombre": "Aseo y vigilancia", "tipo_retefuente": "servicios"},
    {"categoria": "Operaciones", "subcategoria": "Combustible",        "alegra_id": 5498, "cuenta_codigo": "519535", "cuenta_nombre": "Combustibles y lubricantes", "tipo_retefuente": "compras"},
    # MARKETING
    {"categoria": "Marketing", "subcategoria": "Publicidad",  "alegra_id": 5495, "cuenta_codigo": "519520", "cuenta_nombre": "Gastos de representación/Publicidad", "tipo_retefuente": "otros"},
    {"categoria": "Marketing", "subcategoria": "Eventos",     "alegra_id": 5495, "cuenta_codigo": "519520", "cuenta_nombre": "Gastos de representación/Eventos", "tipo_retefuente": "otros"},
    # IMPUESTOS
    {"categoria": "Impuestos", "subcategoria": "ICA",        "alegra_id": 5478, "cuenta_codigo": "511505", "cuenta_nombre": "Industria y Comercio (ICA)", "tipo_retefuente": "impuesto"},
    {"categoria": "Impuestos", "subcategoria": "Predial",    "alegra_id": 5478, "cuenta_codigo": "511505", "cuenta_nombre": "Industria y Comercio (predial)", "tipo_retefuente": "impuesto"},
    # FINANCIERO
    {"categoria": "Financiero", "subcategoria": "Intereses",           "alegra_id": 5533, "cuenta_codigo": "615020", "cuenta_nombre": "Intereses (créditos directos)", "tipo_retefuente": "otros"},
    {"categoria": "Financiero", "subcategoria": "Comisiones_Bancarias", "alegra_id": 5508, "cuenta_codigo": "530515", "cuenta_nombre": "Comisiones bancarias", "tipo_retefuente": "otros"},
    {"categoria": "Financiero", "subcategoria": "Gastos_Bancarios",    "alegra_id": 5507, "cuenta_codigo": "530505", "cuenta_nombre": "Gastos bancarios", "tipo_retefuente": "otros"},
    {"categoria": "Financiero", "subcategoria": "Seguros",             "alegra_id": 5493, "cuenta_codigo": "5195",   "cuenta_nombre": "Gastos generales (seguros)", "tipo_retefuente": "otros"},
    {"categoria": "Financiero", "subcategoria": "GMF",                 "alegra_id": 5509, "cuenta_codigo": "531520", "cuenta_nombre": "Gravamen al movimiento financiero", "tipo_retefuente": "otros"},
    # OTROS
    {"categoria": "Otros", "subcategoria": "Varios",         "alegra_id": 5493, "cuenta_codigo": "5195",   "cuenta_nombre": "Gastos generales", "tipo_retefuente": "otros"},
    {"categoria": "Otros", "subcategoria": "Representacion", "alegra_id": 5495, "cuenta_codigo": "519520", "cuenta_nombre": "Gastos de representación", "tipo_retefuente": "otros"},
    {"categoria": "Otros", "subcategoria": "Depreciacion",   "alegra_id": 5501, "cuenta_codigo": "5160",   "cuenta_nombre": "Depreciación", "tipo_retefuente": None},
]


def _lookup_plan_cuentas_local(categoria: str, subcategoria: str):
    """Lookup account from local PLAN_CUENTAS_RODDOS (no DB needed)."""
    cat_n = _normalize(categoria)
    sub_n = _normalize(subcategoria)
    for entry in PLAN_CUENTAS_RODDOS:
        if _normalize(entry["categoria"]) == cat_n and _normalize(entry["subcategoria"]) == sub_n:
            return entry
    # partial subcategoria match
    for entry in PLAN_CUENTAS_RODDOS:
        if _normalize(entry["categoria"]) == cat_n and sub_n in _normalize(entry["subcategoria"]):
            return entry
    return None

# ReteFuente: cuenta_id (Alegra) + tasa + label
RETEFUENTE_CONFIG = {
    "arrendamiento": {"cuenta_id": 5386, "tasa": 0.035, "label": "Ret. Arriendo 3.5%"},
    "servicios":     {"cuenta_id": 5383, "tasa": 0.04,  "label": "Ret. Servicios 4%"},
    "honorarios_pn": {"cuenta_id": 5381, "tasa": 0.10,  "label": "Ret. Honorarios PN 10%"},
    "honorarios_pj": {"cuenta_id": 5382, "tasa": 0.11,  "label": "Ret. Honorarios PJ 11%"},
    "compras":       {"cuenta_id": 5388, "tasa": 0.025, "label": "Ret. Compras 2.5%"},
    "nomina":        None,
    "impuesto":      None,
    "otros":         None,
}

CUENTA_IVA_SERV    = 5408   # IVA Descontable por Servicios
CUENTA_IVA_COMP    = 5406   # IVA Descontable en Compras 19%
CUENTA_PROVEEDORES = 5376   # Cuentas por pagar a proveedores nacionales
UVT_2026           = 49799
UMBRAL_COMPRAS     = 27 * UVT_2026  # ~1,344,573 para aplicar retefuente compras
UMBRAL_SERVICIOS   = 4  * UVT_2026  # ~199,196   para aplicar retefuente servicios

HEADER_COLOR   = "0F2A5C"
ACCENT_COLOR   = "00C4D4"
EXAMPLE_FILL   = "E8F0FE"

# Job store en memoria (por proceso)
_jobs: dict[str, dict] = {}

# ── helpers ───────────────────────────────────────────────────────────────────

def _normalize(s: str) -> str:
    return unicodedata.normalize("NFD", s.strip().lower()).encode("ascii", "ignore").decode()


def _parse_number(raw) -> float:
    if raw is None:
        return 0.0
    if isinstance(raw, (int, float)):
        return float(raw)
    cleaned = str(raw).replace("$", "").replace("\xa0", "").replace(".", "").replace(",", ".").strip()
    try:
        return float(cleaned)
    except ValueError:
        return 0.0


def _parse_bool(raw) -> bool:
    if raw is None:
        return False
    return str(raw).strip().lower() in ("si", "sí", "yes", "true", "1", "x")


def _calcular_retenciones(row: dict, es_autoretenedor: bool) -> dict:
    """Calcula retenciones para una fila de gasto.
    Usa tipo_retefuente si está presente (del plan_cuentas), o tipo_gasto legacy."""
    # tipo_retefuente comes from plan_cuentas lookup; tipo_gasto is the old column
    tipo_gasto = row.get("tipo_retefuente") or row.get("tipo_gasto", "otros")
    monto_base = row.get("monto_sin_iva", 0.0)

    retefuente_monto = 0.0
    retefuente_cuenta_id = None
    retefuente_label = ""

    if not es_autoretenedor:
        cfg = RETEFUENTE_CONFIG.get(tipo_gasto)
        if cfg:
            tasa = cfg["tasa"]
            # Umbral para compras y servicios
            if tipo_gasto == "compras" and monto_base < UMBRAL_COMPRAS:
                pass  # Debajo del umbral, sin retención
            elif tipo_gasto == "servicios" and monto_base < UMBRAL_SERVICIOS:
                pass  # Debajo del umbral, sin retención
            else:
                retefuente_monto = round(monto_base * tasa)
                retefuente_cuenta_id = cfg["cuenta_id"]
                retefuente_label = cfg["label"]

    iva_monto = 0.0
    iva_cuenta_id = None
    if _parse_bool(row.get("incluye_iva")):
        iva_monto = round(monto_base * 0.19)
        iva_cuenta_id = (
            CUENTA_IVA_SERV if tipo_gasto in ("servicios", "honorarios_pn", "honorarios_pj", "arrendamiento")
            else CUENTA_IVA_COMP
        )

    bruto_total = monto_base + iva_monto
    neto_pagar  = bruto_total - retefuente_monto

    return {
        "iva_monto":             iva_monto,
        "iva_cuenta_id":         iva_cuenta_id,
        "retefuente_monto":      retefuente_monto,
        "retefuente_cuenta_id":  retefuente_cuenta_id,
        "retefuente_label":      retefuente_label,
        "bruto_total":           bruto_total,
        "neto_pagar":            neto_pagar,
    }


HEADER_COLS = [
    "Fecha", "Proveedor", "NIT_Proveedor", "Concepto",
    "Monto_Sin_IVA", "Incluye_IVA", "Categoria", "Subcategoria",
    "Tipo_Persona", "Es_Autoretenedor", "Forma_Pago", "Mes_Periodo", "Notas",
]

ALIASES = {
    "fecha":             ["fecha", "date", "fec"],
    "proveedor":         ["proveedor", "proveedor_nombre", "empresa", "supplier"],
    "nit_proveedor":     ["nit_proveedor", "nit", "cedula", "identification", "cc"],
    "concepto":          ["concepto", "descripcion", "descripcion", "detalle", "description"],
    "monto_sin_iva":     ["monto_sin_iva", "monto", "subtotal", "valor_base", "valor"],
    "incluye_iva":       ["incluye_iva", "incluye iva", "con_iva", "tiene_iva"],
    "categoria":         ["categoria", "categoría", "category", "tipo_gasto", "tipo"],
    "subcategoria":      ["subcategoria", "subcategoría", "subcategory", "sub_categoria"],
    "tipo_persona":      ["tipo_persona", "tipo_proveedor"],
    "es_autoretenedor":  ["es_autoretenedor", "autoretenedor", "auto_retenedor"],
    "forma_pago":        ["forma_pago", "forma de pago", "pago"],
    "mes_periodo":       ["mes_periodo", "periodo", "mes", "period"],
    "notas":             ["notas", "nota", "observaciones", "notes", "referencia"],
}

REQUIRED_FIELDS = ["proveedor", "monto_sin_iva"]

_CATEGORIAS_VALIDAS = sorted(set(e["categoria"] for e in PLAN_CUENTAS_RODDOS))
_SUBCATEGORIAS_VALIDAS = sorted(set(e["subcategoria"] for e in PLAN_CUENTAS_RODDOS))

_INSTRUCT_ROWS = [
    ("Fecha",            "Fecha del gasto YYYY-MM-DD",                    "2026-01-15",       "✅"),
    ("Proveedor",        "Nombre del proveedor o empresa",                "Inmobiliaria XYZ", "✅"),
    ("NIT_Proveedor",    "NIT o cédula del proveedor",                    "900123456-1",      "Opcional"),
    ("Concepto",         "Descripción del gasto",                         "Arriendo oficina enero 2026","✅"),
    ("Monto_Sin_IVA",    "Valor base del gasto, sin IVA, sin puntos",     "3000000",          "✅"),
    ("Incluye_IVA",      "¿El gasto tiene IVA? Si / No",                  "No",               "Opcional"),
    ("Categoria",        " | ".join(_CATEGORIAS_VALIDAS),                 "Operaciones",      "✅"),
    ("Subcategoria",     "Ver hoja Plan_Cuentas para valores válidos",    "Arriendo",         "✅"),
    ("Tipo_Persona",     "PN (persona natural) | PJ (persona jurídica) | Empresa",
                         "PJ",              "Opcional"),
    ("Es_Autoretenedor", "Si / No — los autoretenedores no tienen ReteFuente","No",           "Opcional"),
    ("Forma_Pago",       "Contado | Credito_15 | Credito_30 | Credito_60 | Credito_90",
                         "Contado",         "✅"),
    ("Mes_Periodo",      "Mes contable YYYY-MM",                          "2026-01",          "Opcional"),
    ("Notas",            "Observaciones o referencia del documento",      "Contrato #123",    "Opcional"),
]


# ── GET /gastos/plantilla ──────────────────────────────────────────────────────
@router.get("/plantilla")
async def descargar_plantilla_gastos(current_user=Depends(get_current_user)):
    """Genera y descarga la plantilla Excel oficial para carga masiva de gastos RODDOS."""
    wb = openpyxl.Workbook()

    ex_fill     = PatternFill("solid", fgColor=EXAMPLE_FILL)
    hdr_fill    = PatternFill("solid", fgColor=HEADER_COLOR)
    hdr_font    = Font(bold=True, color="FFFFFF", size=11)
    hdr_align   = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── Hoja 1: Gastos ────────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Gastos"

    # Título superior
    ws1.merge_cells("A1:L1")
    title_cell = ws1.cell(row=1, column=1, value="RODDOS — Plantilla Carga Masiva de Gastos")
    title_cell.fill = PatternFill("solid", fgColor=HEADER_COLOR)
    title_cell.font = Font(bold=True, color="FFFFFF", size=13)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[1].height = 28

    # Sub-header
    ws1.merge_cells("A2:L2")
    sub_cell = ws1.cell(row=2, column=1, value="Completa desde la fila 4. La fila 3 es un ejemplo — será ignorada al cargar.")
    sub_cell.fill = PatternFill("solid", fgColor=ACCENT_COLOR)
    sub_cell.font = Font(bold=False, color="FFFFFF", size=10, italic=True)
    sub_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[2].height = 16

    # Headers row 3 ... wait, let's use row 3 as headers, row 4 as example
    for col_idx, col_name in enumerate(HEADER_COLS, start=1):
        cell = ws1.cell(row=3, column=col_idx, value=col_name)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = hdr_align
        cell.border = border
    ws1.row_dimensions[3].height = 22

    # Example row (row 4) — colored differently
    example_vals = [
        date.today().isoformat(),
        "Inmobiliaria XYZ S.A.S.",
        "900123456-1",
        "Arriendo oficina enero 2026",
        3000000,
        "No",
        "Operaciones",
        "Arriendo",
        "PJ",
        "No",
        "Contado",
        f"{date.today().year}-{date.today().month:02d}",
        "Ejemplo — no modificar esta fila",
    ]
    for col_idx, val in enumerate(example_vals, start=1):
        cell = ws1.cell(row=4, column=col_idx, value=val)
        cell.fill = ex_fill
        cell.border = border
        cell.alignment = Alignment(
            horizontal="center" if col_idx not in (2, 4, 12) else "left",
            vertical="center"
        )
        cell.font = Font(italic=True, color="5C6BC0")

    # Column widths
    col_widths = [13, 26, 16, 30, 14, 12, 16, 16, 14, 16, 14, 13, 24]
    for i, w in enumerate(col_widths, start=1):
        ws1.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # Data validations (start from row 5, the first real data row)
    def _dv(formula: str, col_letter: str, error_msg: str, prompt: str):
        dv = DataValidation(type="list", formula1=formula, allow_blank=True, showDropDown=False)
        dv.error       = error_msg
        dv.errorTitle  = "Valor inválido"
        dv.prompt      = prompt
        dv.promptTitle = "Opciones"
        ws1.add_data_validation(dv)
        dv.sqref = f"{col_letter}5:{col_letter}1000"

    _dv('"Si,No"', "F", "Escribe Si o No", "¿El gasto incluye IVA del 19%?")
    _dv('"' + ",".join(_CATEGORIAS_VALIDAS) + '"',
        "G", "Usa una categoría del plan de cuentas", "Categoría contable — ver hoja Plan_Cuentas")
    _dv('"' + ",".join(_SUBCATEGORIAS_VALIDAS) + '"',
        "H", "Usa una subcategoría del plan de cuentas", "Subcategoría — ver hoja Plan_Cuentas")
    _dv('"PN,PJ,Empresa"', "I", "Usa PN / PJ / Empresa", "Tipo de persona del proveedor")
    _dv('"Si,No"', "J", "Escribe Si o No", "¿Es autoretenedor?")
    _dv('"Contado,Credito_15,Credito_30,Credito_60,Credito_90"',
        "K", "Usa una de las formas de pago listadas", "Contado = asiento | Credito_N = factura")

    # Freeze pane below headers + example
    ws1.freeze_panes = "A5"

    # ── Hoja 3: Plan_Cuentas ─────────────────────────────────────────────────
    ws3 = wb.create_sheet("Plan_Cuentas")
    ws3.merge_cells("A1:F1")
    pc_title = ws3.cell(row=1, column=1, value="PLAN DE CUENTAS RODDOS — Mapeo a Alegra (IDs reales)")
    pc_title.fill = PatternFill("solid", fgColor=HEADER_COLOR)
    pc_title.font = Font(bold=True, color="FFFFFF", size=12)
    pc_title.alignment = Alignment(horizontal="center", vertical="center")
    ws3.row_dimensions[1].height = 24

    pc_hdrs = ["Categoria", "Subcategoria", "Cuenta Alegra", "Código PUC", "Alegra ID", "Usar en template"]
    for ci, h in enumerate(pc_hdrs, start=1):
        c = ws3.cell(row=2, column=ci, value=h)
        c.fill = PatternFill("solid", fgColor="D9E1F2")
        c.font = Font(bold=True)
        c.border = border

    for ri, entry in enumerate(PLAN_CUENTAS_RODDOS, start=3):
        ws3.cell(row=ri, column=1, value=entry["categoria"]).font = Font(bold=True)
        ws3.cell(row=ri, column=2, value=entry["subcategoria"])
        ws3.cell(row=ri, column=3, value=entry["cuenta_nombre"])
        ws3.cell(row=ri, column=4, value=entry["cuenta_codigo"])
        ws3.cell(row=ri, column=5, value=entry["alegra_id"]).font = Font(bold=True, color="1F497D")
        ws3.cell(row=ri, column=6, value=f"Categoria: {entry['categoria']} / Subcategoria: {entry['subcategoria']}")
        for ci in range(1, 7):
            ws3.cell(row=ri, column=ci).border = border
        ws3.row_dimensions[ri].height = 15

    ws3.column_dimensions["A"].width = 15
    ws3.column_dimensions["B"].width = 22
    ws3.column_dimensions["C"].width = 36
    ws3.column_dimensions["D"].width = 14
    ws3.column_dimensions["E"].width = 12
    ws3.column_dimensions["F"].width = 44

    # ── Hoja 2: Instrucciones ─────────────────────────────────────────────────
    ws2 = wb.create_sheet("Instrucciones")

    ws2.merge_cells("A1:D1")
    t = ws2.cell(row=1, column=1, value="INSTRUCCIONES — Carga Masiva de Gastos RODDOS")
    t.font = Font(bold=True, color=HEADER_COLOR, size=14)
    t.alignment = Alignment(horizontal="left", vertical="center")
    ws2.row_dimensions[1].height = 30

    ws2.merge_cells("A2:D2")
    ws2.cell(row=2, column=1,
             value="⚠️ La fila 4 (azul claro) es un EJEMPLO. El sistema la ignora al procesar.")
    ws2.cell(row=2, column=1).font = Font(bold=True, color="C62828", size=11)
    ws2.row_dimensions[2].height = 18

    hdr_labels = ["Campo", "Descripción y valores válidos", "Ejemplo", "Req."]
    for ci, lbl in enumerate(hdr_labels, start=1):
        c = ws2.cell(row=4, column=ci, value=lbl)
        c.fill = PatternFill("solid", fgColor="D9E1F2")
        c.font = Font(bold=True)
        c.border = border
        c.alignment = hdr_align
    ws2.row_dimensions[4].height = 18

    for ri, (campo, desc, ej, oblig) in enumerate(_INSTRUCT_ROWS, start=5):
        ws2.cell(row=ri, column=1, value=campo).font = Font(bold=True)
        ws2.cell(row=ri, column=2, value=desc)
        ws2.cell(row=ri, column=3, value=ej).font = Font(italic=True, color="1F497D")
        ws2.cell(row=ri, column=4, value=oblig).alignment = Alignment(horizontal="center")
        for ci in range(1, 5):
            ws2.cell(row=ri, column=ci).border = border
        ws2.row_dimensions[ri].height = 16

    ws2.column_dimensions["A"].width = 18
    ws2.column_dimensions["B"].width = 58
    ws2.column_dimensions["C"].width = 26
    ws2.column_dimensions["D"].width = 8

    # Notes
    ns = len(_INSTRUCT_ROWS) + 7
    notas = [
        "NOTAS IMPORTANTES:",
        "• No modificar los encabezados de la fila 3 en la hoja Gastos",
        "• La fila 4 (ejemplo) es ignorada automáticamente por el sistema",
        "• Auteco Kawasaki (NIT 860024781) ya está marcado como autoretenedor — sin ReteFuente",
        "• Forma_Pago = Contado → crea asiento contable (journal entry) en Alegra",
        "• Forma_Pago = Credito_30 → crea factura de compra con vencimiento a 30 días",
        "• Guardar como .xlsx antes de subir al sistema",
        "• Máximo recomendado: 200 filas por archivo",
    ]
    for i, nota in enumerate(notas):
        c = ws2.cell(row=ns + i, column=1, value=nota)
        c.alignment = Alignment(wrap_text=True)
        if i == 0:
            c.font = Font(bold=True, color=HEADER_COLOR)
        ws2.merge_cells(f"A{ns + i}:D{ns + i}")
        ws2.row_dimensions[ns + i].height = 16

    # Serialize
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="RODDOS_Plantilla_Gastos.xlsx"'},
    )


# ── POST /gastos/cargar ────────────────────────────────────────────────────────
@router.post("/cargar")
async def cargar_gastos_excel(
    file: UploadFile = File(...),
    current_user=Depends(get_current_user),
):
    """Recibe Excel, parsea, valida y devuelve preview con retenciones calculadas."""
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="El archivo debe ser .xlsx o .xls")

    contents = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
        ws = wb.active
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Error al leer el archivo: {e}")

    # Find header row (row 1, 2 or 3)
    header_row_num = None
    col_map: dict[str, int] = {}
    for rn in [1, 2, 3]:
        raw_row = [str(c.value or "").strip() for c in ws[rn]]
        norm_row = [_normalize(h) for h in raw_row]
        temp_map: dict[str, int] = {}
        # Build column map — 2-pass: exact match first, then substring
        for field, aliases in ALIASES.items():
            naliases = [_normalize(a) for a in aliases]
            # Pass 1: exact match
            for i, h in enumerate(norm_row):
                if not h:
                    continue
                if h in naliases:
                    temp_map[field] = i
                    break
            if field in temp_map:
                continue
            # Pass 2: substring match (min len 4 to avoid false positives like "iva" in "monto_sin_iva")
            for i, h in enumerate(norm_row):
                if not h or len(h) < 3:
                    continue
                if any((len(a) >= 4 and a in h) or (len(h) >= 4 and h in a) for a in naliases):
                    temp_map.setdefault(field, i)
                    break
        if len(temp_map) >= 3:
            header_row_num = rn
            col_map = temp_map
            break

    if not col_map:
        return {
            "ok": False,
            "error": "No se encontraron columnas válidas. Descarga la plantilla oficial y llena tus datos.",
            "sugerencia": "Usa el botón 'Descargar Plantilla' para obtener el formato correcto.",
        }

    missing = [f for f in REQUIRED_FIELDS if f not in col_map]
    if missing:
        return {
            "ok": False,
            "error": f"Columnas requeridas no encontradas: {missing}. Descarga la plantilla oficial.",
            "columnas_detectadas": [str(c.value or "") for c in ws[header_row_num]],
        }

    # Load autoretenedores list
    autoretenedores_db = await db.proveedores_config.find(
        {"es_autoretenedor": True}, {"_id": 0, "nombre": 1, "nit": 1}
    ).to_list(100)
    auto_nombres = {a["nombre"].lower() for a in autoretenedores_db}
    auto_nits    = {str(a.get("nit", "")).replace("-", "") for a in autoretenedores_db if a.get("nit")}

    # Parse rows — skip header row and example row
    gastos_preview = []
    row_errors: list[str] = []
    first_data_row = True
    example_proveedor = "auteco kawasaki"

    start_row = (header_row_num or 3) + 1
    for row_num, row in enumerate(ws.iter_rows(min_row=start_row, values_only=True), start=start_row):
        if all(v is None for v in row):
            continue

        def _get(field):
            idx = col_map.get(field)
            if idx is not None and idx < len(row):
                return row[idx]
            return None

        proveedor = str(_get("proveedor") or "").strip()
        if not proveedor:
            continue

        # Skip example row
        if first_data_row and example_proveedor in proveedor.lower():
            first_data_row = False
            continue
        first_data_row = False

        monto = _parse_number(_get("monto_sin_iva"))
        if monto <= 0:
            row_errors.append(f"Fila {row_num}: Monto_Sin_IVA inválido para '{proveedor}' — omitida")
            continue

        nit          = str(_get("nit_proveedor") or "").strip()
        concepto     = str(_get("concepto") or "").strip() or f"Gasto {proveedor}"
        incluye_iva  = str(_get("incluye_iva") or "").strip()
        # Support both old tipo_gasto column and new categoria/subcategoria
        raw_categoria    = _normalize(str(_get("categoria") or "").strip())
        raw_subcategoria = _normalize(str(_get("subcategoria") or "").strip())
        tipo_gasto   = raw_categoria  # used as fallback for legacy support
        tipo_persona = str(_get("tipo_persona") or "PJ").strip().upper()
        es_auto_col  = _parse_bool(_get("es_autoretenedor"))
        forma_pago   = str(_get("forma_pago") or "Contado").strip()
        mes_periodo  = str(_get("mes_periodo") or "").strip()
        notas        = str(_get("notas") or "").strip()

        # Normalize tipo_gasto (backward compat: 'arrendamiento' etc.)
        tipo_map = {
            "arriendo": "arrendamiento", "rent": "arrendamiento",
            "servicio": "servicios", "service": "servicios",
            "honorarios": "honorarios_pn",  # fallback
            "salario": "nomina", "sueldo": "nomina",
            "impuesto": "impuesto", "tax": "impuesto",
        }
        tipo_gasto = tipo_map.get(tipo_gasto, tipo_gasto)
        if tipo_gasto not in CUENTAS_GASTO:
            tipo_gasto = "otros"

        # Determine tipo_persona for honorarios
        if tipo_gasto == "honorarios_pn" and tipo_persona in ("PJ", "EMPRESA"):
            tipo_gasto = "honorarios_pj"
        elif tipo_gasto in ("honorarios_pj",) and tipo_persona == "PN":
            tipo_gasto = "honorarios_pn"

        # Resolve account from plan_cuentas (new) or tipo_gasto (legacy)
        plan_entry = None
        if raw_categoria and raw_subcategoria:
            plan_entry = _lookup_plan_cuentas_local(raw_categoria, raw_subcategoria)

        if plan_entry:
            cuenta_gasto_id    = plan_entry["alegra_id"]
            cuenta_gasto_nombre = plan_entry["cuenta_nombre"]
            tipo_retefuente    = plan_entry.get("tipo_retefuente") or "otros"
        else:
            cuenta_gasto_id    = CUENTAS_GASTO[tipo_gasto]["id"]
            cuenta_gasto_nombre = CUENTAS_GASTO[tipo_gasto]["nombre"]
            tipo_retefuente    = tipo_gasto

        # Check autoretenedor: column OR DB list
        nit_clean = nit.replace("-", "").replace(" ", "")
        es_autoretenedor = es_auto_col or (
            proveedor.lower() in auto_nombres or nit_clean in auto_nits
        )

        # Fecha
        fecha_raw = _get("fecha")
        if isinstance(fecha_raw, datetime):
            fecha = fecha_raw.date().isoformat()
        elif fecha_raw:
            fecha = str(fecha_raw)[:10]
        else:
            fecha = date.today().isoformat()

        row_dict = {
            "id":               str(uuid.uuid4()),
            "row_num":          row_num,
            "fecha":            fecha,
            "proveedor":        proveedor,
            "nit_proveedor":    nit,
            "concepto":         concepto,
            "monto_sin_iva":    monto,
            "incluye_iva":      str(incluye_iva).strip(),
            "categoria":        raw_categoria or tipo_gasto,
            "subcategoria":     raw_subcategoria or "",
            "tipo_gasto":       tipo_gasto,
            "tipo_retefuente":  tipo_retefuente,
            "tipo_persona":     tipo_persona,
            "es_autoretenedor": es_autoretenedor,
            "forma_pago":       forma_pago,
            "mes_periodo":      mes_periodo,
            "notas":            notas,
        }

        reten = _calcular_retenciones(row_dict, es_autoretenedor)
        row_dict.update(reten)
        row_dict["cuenta_gasto_id"]     = cuenta_gasto_id
        row_dict["cuenta_gasto_nombre"] = cuenta_gasto_nombre

        gastos_preview.append(row_dict)

    if not gastos_preview:
        msg = "El archivo no tiene filas con datos válidos."
        if row_errors:
            msg += f" Errores encontrados: {row_errors[0]}"
        return {"ok": False, "error": msg}

    total_monto    = sum(g["monto_sin_iva"] for g in gastos_preview)
    total_iva      = sum(g["iva_monto"] for g in gastos_preview)
    total_retefuen = sum(g["retefuente_monto"] for g in gastos_preview)
    total_neto     = sum(g["neto_pagar"] for g in gastos_preview)
    contado_count  = sum(1 for g in gastos_preview if g["forma_pago"].lower() == "contado")
    credito_count  = len(gastos_preview) - contado_count

    return {
        "ok":           True,
        "gastos":       gastos_preview,
        "advertencias": row_errors,
        "resumen": {
            "total_filas":      len(gastos_preview),
            "total_monto_base": total_monto,
            "total_iva":        total_iva,
            "total_retefuente": total_retefuen,
            "total_neto_pagar": total_neto,
            "contado":          contado_count,
            "credito":          credito_count,
        },
    }


# ── Lógica de procesamiento individual ────────────────────────────────────────

async def _find_or_create_contact(alegra: AlegraService, proveedor: str, nit: str, tipo_persona: str) -> Optional[int]:
    """Busca o crea un contacto en Alegra. Retorna el ID del contacto."""
    try:
        contacts = await alegra.request("contacts", params={"name": proveedor[:30]})
        if isinstance(contacts, list):
            nit_clean = nit.replace("-", "").replace(" ", "")
            for c in contacts:
                c_nit = str(c.get("identification", "")).replace("-", "").replace(" ", "")
                if c_nit and c_nit == nit_clean:
                    return c.get("id")
                if c.get("name", "").lower() == proveedor.lower():
                    return c.get("id")
    except Exception:
        pass

    # Create contact
    try:
        name_parts = proveedor.strip().split(" ", 1)
        first = name_parts[0]
        last  = name_parts[1] if len(name_parts) > 1 else "."
        tipo  = ["provider"]
        contact_payload = {
            "name":           proveedor,
            "nameObject":     {"firstName": first, "lastName": last},
            "identification": nit or proveedor[:10],
            "type":           tipo,
        }
        result = await alegra.request("contacts", "POST", contact_payload)
        if isinstance(result, dict) and result.get("id"):
            return result["id"]
    except Exception as e:
        logger.warning("No se pudo crear contacto %s: %s", proveedor, e)
    return None


async def _process_row(alegra: AlegraService, gasto: dict) -> dict:
    """Procesa una fila: crea journal-entry o bill en Alegra. Retorna {'ok', 'id', 'error'}."""
    forma_pago = gasto.get("forma_pago", "Contado")
    is_contado = forma_pago.lower() == "contado"

    cuenta_gasto_id    = gasto.get("cuenta_gasto_id", 5495)
    iva_monto          = gasto.get("iva_monto", 0.0)
    iva_cuenta_id      = gasto.get("iva_cuenta_id")
    retefuente_monto   = gasto.get("retefuente_monto", 0.0)
    retefuente_cuenta  = gasto.get("retefuente_cuenta_id")
    bruto_total        = gasto.get("bruto_total", gasto["monto_sin_iva"])
    neto_pagar         = gasto.get("neto_pagar", bruto_total)

    obs = (
        f"Gasto masivo — {gasto['concepto']} "
        f"({gasto['proveedor']}"
        + (f", NIT {gasto['nit_proveedor']}" if gasto.get('nit_proveedor') else "")
        + ")"
        + (f" · {gasto['mes_periodo']}" if gasto.get('mes_periodo') else "")
        + (f" · {gasto['notas']}" if gasto.get('notas') else "")
    ).strip()

    try:
        if is_contado:
            # Journal entry (causación contable)
            entries = []
            entries.append({"id": cuenta_gasto_id, "debit": round(gasto["monto_sin_iva"]), "credit": 0})
            if iva_monto > 0 and iva_cuenta_id:
                entries.append({"id": iva_cuenta_id, "debit": round(iva_monto), "credit": 0})
            if retefuente_monto > 0 and retefuente_cuenta:
                entries.append({"id": retefuente_cuenta, "debit": 0, "credit": round(retefuente_monto)})
            entries.append({"id": CUENTA_PROVEEDORES, "debit": 0, "credit": round(neto_pagar)})

            payload = {
                "date":         gasto["fecha"],
                "observations": obs,
                "entries":      entries,
            }
            result = await alegra.request("journals", "POST", payload)
            doc_id = str(result.get("id", "")) if isinstance(result, dict) else ""
            return {"ok": True, "id": doc_id, "tipo": "journal-entry", "numero": result.get("number", "")}

        else:
            # Bill (factura de compra a crédito)
            days = 30
            try:
                days = int(forma_pago.split("_")[1])
            except Exception:
                pass
            fecha_base = date.fromisoformat(gasto["fecha"]) if gasto.get("fecha") else date.today()
            due_date   = (fecha_base + timedelta(days=days)).isoformat()

            # Find/create contact
            contact_id = await _find_or_create_contact(
                alegra, gasto["proveedor"], gasto.get("nit_proveedor", ""), gasto.get("tipo_persona", "PJ")
            )

            bill_payload = {
                "date":        gasto["fecha"],
                "dueDate":     due_date,
                "observations": obs,
                "purchases": {
                    "items": [
                        {
                            "description": gasto["concepto"],
                            "price":       round(bruto_total),
                            "quantity":    1,
                        }
                    ]
                },
            }
            if contact_id:
                bill_payload["contact"] = {"id": contact_id}

            result = await alegra.request("bills", "POST", bill_payload)
            doc_id = str(result.get("id", "")) if isinstance(result, dict) else ""
            return {"ok": True, "id": doc_id, "tipo": "bill", "vencimiento": due_date, "numero": result.get("number", "")}

    except Exception as e:
        return {"ok": False, "error": str(e)}


async def _run_job(job_id: str, gastos: list[dict], user_email: str):
    """Background task: procesa todas las filas y actualiza el job."""
    alegra = AlegraService(db)
    total  = len(gastos)
    exitosos: list[dict] = []
    errores:  list[dict] = []

    _jobs[job_id].update({"estado": "procesando", "total": total, "procesados": 0})

    for i, gasto in enumerate(gastos):
        result = await _process_row(alegra, gasto)
        gasto_out = {
            "row_num":   gasto.get("row_num", i + 1),
            "proveedor": gasto["proveedor"],
            "concepto":  gasto["concepto"],
            "monto":     gasto["monto_sin_iva"],
            "fecha":     gasto["fecha"],
            "forma_pago": gasto.get("forma_pago", "Contado"),
        }
        if result["ok"]:
            gasto_out.update({"alegra_id": result["id"], "tipo": result["tipo"], "numero": result.get("numero", "")})
            exitosos.append(gasto_out)
        else:
            gasto_out["error"] = result["error"]
            errores.append(gasto_out)

        _jobs[job_id]["procesados"] = i + 1
        _jobs[job_id]["exitosos"]   = len(exitosos)
        _jobs[job_id]["errores"]    = len(errores)

    # Save to DB — incluye alegra_ids de los exitosos para poder hacer cleanup después
    now = datetime.now(timezone.utc).isoformat()
    exitosos_ids = [
        {"alegra_id": g.get("alegra_id",""), "tipo": g.get("tipo",""), "fecha": g.get("fecha",""), "proveedor": g.get("proveedor",""), "monto": g.get("monto",0)}
        for g in exitosos if g.get("alegra_id")
    ]
    await db.roddos_events.insert_one({
        "id":           job_id,
        "event_type":   "gasto.masivo.registrado",
        "procesados":   len(exitosos) + len(errores),
        "exitosos":     len(exitosos),
        "errores":      len(errores),
        "alegra_ids":   exitosos_ids,   # IDs reales de Alegra para cleanup futuro
        "filas_error":  errores,
        "creado_por":   user_email,
        "fecha":        now,
    })

    _jobs[job_id].update({
        "estado":       "completado",
        "exitosos":     len(exitosos),
        "errores_lista":  errores,
        "completado_en":  now,
    })


# ── POST /gastos/procesar ─────────────────────────────────────────────────────
class ProcesarGastosReq(BaseModel):
    gastos: List[dict]


@router.post("/procesar")
async def procesar_gastos(
    req: ProcesarGastosReq,
    background_tasks: BackgroundTasks,
    current_user=Depends(get_current_user),
):
    """Inicia procesamiento masivo (siempre asíncrono). Retorna job_id."""
    if not req.gastos:
        raise HTTPException(status_code=400, detail="No hay gastos para procesar")

    job_id = str(uuid.uuid4())
    _jobs[job_id] = {
        "job_id":     job_id,
        "estado":     "iniciando",
        "total":      len(req.gastos),
        "procesados": 0,
        "exitosos":   0,
        "errores":    0,
        "creado_en":  datetime.now(timezone.utc).isoformat(),
    }

    background_tasks.add_task(
        _run_job, job_id, req.gastos, current_user.get("email", "")
    )

    return {"ok": True, "job_id": job_id, "total": len(req.gastos)}


async def _invalidate_cfo_after_bulk():
    """Called in background after bulk gastos job completes to refresh CFO cache."""
    try:
        from routers.cfo import invalidar_cache_cfo
        await invalidar_cache_cfo()
    except Exception:
        pass


# ── GET /gastos/jobs/{job_id} ──────────────────────────────────────────────────
@router.get("/jobs/{job_id}")
async def get_job_status(job_id: str, current_user=Depends(get_current_user)):
    """Consulta el estado de un job de carga masiva."""
    job = _jobs.get(job_id)
    if not job:
        # Try DB (job may have been created in a different process restart)
        ev = await db.roddos_events.find_one({"id": job_id, "event_type": "gasto.masivo.registrado"}, {"_id": 0})
        if ev:
            return {
                "job_id":   job_id,
                "estado":   "completado",
                "total":    ev.get("procesados", 0),
                "exitosos": ev.get("exitosos", 0),
                "errores":  ev.get("errores", 0),
                "errores_lista": ev.get("filas_error", []),
            }
        raise HTTPException(status_code=404, detail="Job no encontrado")
    return {k: v for k, v in job.items() if k != "errores_lista"} | {
        "tiene_errores": len(job.get("errores_lista", [])) > 0
    }


# ── GET /gastos/reporte-errores/{job_id} ──────────────────────────────────────
@router.get("/reporte-errores/{job_id}")
async def descargar_reporte_errores(job_id: str, current_user=Depends(get_current_user)):
    """Descarga Excel con filas que fallaron en el procesamiento."""
    job = _jobs.get(job_id)
    errores = []

    if job:
        errores = job.get("errores_lista", [])
    else:
        ev = await db.roddos_events.find_one({"id": job_id, "event_type": "gasto.masivo.registrado"}, {"_id": 0})
        if ev:
            errores = ev.get("filas_error", [])

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Errores"

    hdr_fill = PatternFill("solid", fgColor="C62828")
    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    hdr_cols = ["Fila", "Proveedor", "Concepto", "Monto", "Fecha", "Forma_Pago", "Error"]

    for ci, h in enumerate(hdr_cols, start=1):
        c = ws.cell(row=1, column=ci, value=h)
        c.fill = hdr_fill
        c.font = hdr_font
        c.alignment = Alignment(horizontal="center")

    for ri, err in enumerate(errores, start=2):
        ws.cell(row=ri, column=1, value=err.get("row_num", ri - 1))
        ws.cell(row=ri, column=2, value=err.get("proveedor", ""))
        ws.cell(row=ri, column=3, value=err.get("concepto", ""))
        ws.cell(row=ri, column=4, value=err.get("monto", 0))
        ws.cell(row=ri, column=5, value=err.get("fecha", ""))
        ws.cell(row=ri, column=6, value=err.get("forma_pago", ""))
        ws.cell(row=ri, column=7, value=err.get("error", ""))
        ws.cell(row=ri, column=7).font = Font(color="C62828")

    for ci, w in enumerate([8, 26, 30, 14, 13, 14, 50], start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="RODDOS_Gastos_Errores_{job_id[:8]}.xlsx"'},
    )



# ── GET /gastos/journals-creados ─────────────────────────────────────────────
@router.get("/journals-creados")
async def get_journals_creados(
    fecha_desde: str = "",
    fecha_hasta: str = "",
    current_user=Depends(get_current_user),
):
    """Retorna los IDs de Alegra de journals creados por carga masiva, desde MongoDB.
    Útil para cleanup sin tener que llamar a Alegra (evita rate limiting)."""
    query = {"event_type": "gasto.masivo.registrado", "alegra_ids": {"$exists": True, "$ne": []}}
    if fecha_desde or fecha_hasta:
        query["fecha"] = {}
        if fecha_desde:
            query["fecha"]["$gte"] = fecha_desde
        if fecha_hasta:
            query["fecha"]["$lte"] = fecha_hasta + "Z"

    events = await db.roddos_events.find(query, {"_id": 0, "alegra_ids": 1, "fecha": 1, "procesados": 1}).to_list(50)
    all_ids = []
    for evt in events:
        all_ids.extend(evt.get("alegra_ids", []))

    return {
        "total": len(all_ids),
        "eventos": len(events),
        "alegra_ids": all_ids,
    }


@router.get("/plan-cuentas")
async def get_plan_cuentas(current_user=Depends(get_current_user)):
    """Devuelve el plan de cuentas RODDOS con IDs reales de Alegra."""
    return {
        "total": len(PLAN_CUENTAS_RODDOS),
        "plan": [
            {k: v for k, v in entry.items() if k != "tipo_retefuente"}
            for entry in PLAN_CUENTAS_RODDOS
        ],
    }


# ── GET /gastos/cleanup-preview ───────────────────────────────────────────────
class CleanupPreviewReq(BaseModel):
    cuenta_id: int = 5495
    fecha_desde: str = ""
    fecha_hasta: str = ""


# ── POST /gastos/cleanup-preview ─────────────────────────────────────────────
# Background task version — stores results in MongoDB to avoid proxy timeout
async def _run_cleanup_preview(job_id: str, cuenta_id: int, fecha_desde: str, fecha_hasta: str):
    """Pagina Alegra journals en background y filtra los de cuenta incorrecta.
    Usa delays amplios (5s) para evitar rate limiting de Alegra."""
    import asyncio
    alegra = AlegraService(db)
    journals_to_delete = []
    start = 0
    batch_size = 10
    total_revisados = 0

    for page_num in range(30):  # max 300 journals
        retries = 0
        batch = None
        while retries < 4:
            try:
                batch = await alegra.request("journals", params={"limit": batch_size, "start": start})
                break  # success
            except Exception as e:
                retries += 1
                wait = min(5 * retries, 20)  # 5s, 10s, 15s, 20s
                logger.warning("cleanup_preview page=%s retry=%s/%s error=%s — waiting %ss", page_num, retries, 4, e, wait)
                if retries >= 4:
                    batch = None
                    break
                await asyncio.sleep(wait)

        if batch is None or not isinstance(batch, list) or not batch:
            break

        total_revisados += len(batch)
        for j in batch:
            j_date = j.get("date", "")
            if fecha_desde and j_date < fecha_desde:
                continue
            if fecha_hasta and j_date > fecha_hasta:
                continue
            account_ids = [str(e.get("id", "")) for e in j.get("entries", [])]
            if str(cuenta_id) in account_ids:
                total_debito = sum(
                    float(e.get("debit", 0)) for e in j.get("entries", [])
                    if float(e.get("debit", 0)) > 0 and str(e.get("id")) == str(cuenta_id)
                )
                journals_to_delete.append({
                    "alegra_id":        str(j.get("id")),
                    "date":             j_date,
                    "observations":     j.get("observations", j.get("description", "")),
                    "total_debito":     total_debito,
                    "cuenta_incorrecta": next(
                        (e.get("name") for e in j.get("entries", []) if str(e.get("id")) == str(cuenta_id)), "?"
                    ),
                })

        # Update progress in MongoDB after each batch
        await db.gastos_cleanup_jobs.update_one(
            {"job_id": job_id},
            {"$set": {
                "total_revisados": total_revisados,
                "journals_encontrados_hasta_ahora": len(journals_to_delete),
            }},
        )

        start += len(batch)
        if len(batch) < batch_size:
            break
        await asyncio.sleep(5)  # 5s delay between batches — critical for Alegra rate limit

    await db.gastos_cleanup_jobs.update_one(
        {"job_id": job_id},
        {"$set": {
            "estado":                          "completado",
            "total_revisados":                 total_revisados,
            "journals_con_cuenta_incorrecta":  len(journals_to_delete),
            "preview":                         journals_to_delete,
            "alegra_ids":                      [j["alegra_id"] for j in journals_to_delete],
            "fin":                             datetime.now(timezone.utc).isoformat(),
        }},
    )
    logger.info("cleanup_preview completado: %d revisados, %d con cuenta %d", total_revisados, len(journals_to_delete), cuenta_id)

    await db.gastos_cleanup_jobs.update_one(
        {"job_id": job_id},
        {"$set": {
            "estado": "completado",
            "total_revisados": total_revisados,
            "journals_con_cuenta_incorrecta": len(journals_to_delete),
            "preview": journals_to_delete,
            "fin": datetime.now(timezone.utc).isoformat(),
        }},
    )


@router.post("/cleanup-preview")
async def cleanup_journals_preview(
    req: CleanupPreviewReq,
    background_tasks: BackgroundTasks,
    current_user=Depends(get_current_user),
):
    """Inicia búsqueda en background de journals con cuenta incorrecta.
    Usa GET /gastos/cleanup-status/{job_id} para obtener resultados."""
    job_id = str(uuid.uuid4())
    await db.gastos_cleanup_jobs.insert_one({
        "job_id": job_id,
        "estado": "en_progreso",
        "cuenta_id": req.cuenta_id,
        "fecha_desde": req.fecha_desde,
        "fecha_hasta": req.fecha_hasta,
        "inicio": datetime.now(timezone.utc).isoformat(),
    })
    background_tasks.add_task(_run_cleanup_preview, job_id, req.cuenta_id, req.fecha_desde, req.fecha_hasta)
    return {"job_id": job_id, "estado": "en_progreso"}


@router.get("/cleanup-status/{job_id}")
async def cleanup_status(job_id: str, current_user=Depends(get_current_user)):
    """Estado y resultados del cleanup-preview job."""
    job = await db.gastos_cleanup_jobs.find_one({"job_id": job_id}, {"_id": 0})
    if not job:
        raise HTTPException(status_code=404, detail="Job no encontrado")
    return job


# ── POST /gastos/cleanup-execute ─────────────────────────────────────────────
class CleanupExecuteReq(BaseModel):
    alegra_ids: List[str]


async def _run_cleanup_execute(job_id: str, alegra_ids: list, user_email: str):
    """Background task: elimina journals de Alegra con delays anti-rate-limit."""
    import asyncio
    alegra = AlegraService(db)
    eliminados: list[str] = []
    errores_delete: list[dict] = []

    for i, journal_id in enumerate(alegra_ids):
        intentos = 0
        while intentos < 3:
            try:
                await alegra.request(f"journals/{journal_id}", method="DELETE")
                eliminados.append(str(journal_id))
                break
            except Exception as e:
                intentos += 1
                err_msg = str(e)
                if intentos >= 3:
                    errores_delete.append({"id": str(journal_id), "error": err_msg})
                    logger.warning("cleanup_execute: no se pudo eliminar journal %s: %s", journal_id, err_msg)
                else:
                    await asyncio.sleep(3 * intentos)

        # Delay every 10 items to avoid Alegra rate limiting
        if (i + 1) % 10 == 0:
            await asyncio.sleep(1.5)
            # Update progress
            await db.gastos_cleanup_jobs.update_one(
                {"job_id": job_id},
                {"$set": {"eliminados": len(eliminados), "errores": len(errores_delete), "procesados": i + 1}},
            )

    fin = datetime.now(timezone.utc).isoformat()

    # Update final state in MongoDB
    await db.gastos_cleanup_jobs.update_one(
        {"job_id": job_id},
        {"$set": {
            "estado":           "completado",
            "eliminados":       len(eliminados),
            "errores":          len(errores_delete),
            "ids_eliminados":   eliminados,
            "detalle_errores":  errores_delete,
            "procesados":       len(alegra_ids),
            "fin":              fin,
        }},
    )

    # Auditable event
    await db.roddos_events.insert_one({
        "event_type":       "gasto.cleanup.journals",
        "job_id":           job_id,
        "eliminados":       len(eliminados),
        "errores":          len(errores_delete),
        "ids_eliminados":   eliminados,
        "detalle_errores":  errores_delete,
        "ejecutado_por":    user_email,
        "fecha":            fin,
    })
    logger.info("cleanup_execute completado: %d eliminados, %d errores", len(eliminados), len(errores_delete))


@router.post("/cleanup-execute")
async def cleanup_journals_execute(
    req: CleanupExecuteReq,
    background_tasks: BackgroundTasks,
    current_user=Depends(get_current_user),
):
    """Inicia eliminación de journals en background. Retorna job_id inmediatamente.
    Consultar estado en GET /gastos/cleanup-status/{job_id}."""
    if not req.alegra_ids:
        raise HTTPException(status_code=400, detail="No se proporcionaron IDs para eliminar")

    if len(req.alegra_ids) > 200:
        raise HTTPException(status_code=400, detail="Máximo 200 journals por solicitud de limpieza")

    job_id = str(uuid.uuid4())
    await db.gastos_cleanup_jobs.insert_one({
        "job_id":        job_id,
        "tipo":          "execute",
        "estado":        "en_progreso",
        "total":         len(req.alegra_ids),
        "procesados":    0,
        "eliminados":    0,
        "errores":       0,
        "ids_recibidos": list(req.alegra_ids),
        "inicio":        datetime.now(timezone.utc).isoformat(),
    })

    background_tasks.add_task(
        _run_cleanup_execute, job_id, list(req.alegra_ids), current_user.get("email", "")
    )

    return {
        "job_id":           job_id,
        "estado":           "en_progreso",
        "total_a_eliminar": len(req.alegra_ids),
        "mensaje":          f"Eliminación iniciada para {len(req.alegra_ids)} journals. Consulta el progreso en GET /api/gastos/cleanup-status/{job_id}",
    }
