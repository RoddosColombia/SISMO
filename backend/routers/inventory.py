"""Inventory router — Auteco motos (upload, CRUD, register in Alegra, sell)."""
from datetime import datetime, timezone
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from pydantic import BaseModel

from alegra_service import AlegraService
from inventory_service import extract_motos_from_pdf, register_moto_in_alegra
from database import db
from dependencies import get_current_user, require_admin, log_action

router = APIRouter(prefix="/inventario", tags=["inventory"])


@router.post("/upload-pdf")
async def upload_pdf(file: UploadFile = File(...), current_user=Depends(get_current_user)):
    if not file.filename.lower().endswith(".pdf"):
        raise HTTPException(status_code=400, detail="Solo se admiten archivos PDF")
    content = await file.read()
    if len(content) > 20 * 1024 * 1024:
        raise HTTPException(status_code=413, detail="El archivo no puede superar 20MB")
    try:
        motos = await extract_motos_from_pdf(content, file.filename)
    except ValueError as e:
        raise HTTPException(status_code=422, detail=str(e))

    if not motos:
        raise HTTPException(status_code=422, detail="No se encontraron motos en el PDF")

    inserted = []
    for m in motos:
        chasis = m.get("chasis")
        if chasis:
            existing = await db.inventario_motos.find_one({"chasis": chasis}, {"_id": 0})
            if existing:
                continue
        await db.inventario_motos.insert_one({k: v for k, v in m.items()})
        inserted.append(m)

    await log_action(current_user, "/inventario/upload-pdf", "POST", {"filename": file.filename, "motos_found": len(motos)})
    return {"inserted": len(inserted), "total_found": len(motos), "motos": inserted}


@router.get("/motos")
async def get_inventario(
    estado: Optional[str] = None,
    marca: Optional[str] = None,
    current_user=Depends(get_current_user),
):
    query = {}
    if estado:
        query["estado"] = estado
    if marca:
        query["marca"] = {"$regex": marca, "$options": "i"}
    motos = await db.inventario_motos.find(query, {"_id": 0}).sort("created_at", -1).to_list(500)
    return motos


@router.get("/stats")
async def get_inventario_stats(current_user=Depends(get_current_user)):
    total = await db.inventario_motos.count_documents({})
    disponibles      = await db.inventario_motos.count_documents({"estado": "Disponible"})
    vendidas         = await db.inventario_motos.count_documents({"estado": "Vendida"})
    entregadas       = await db.inventario_motos.count_documents({"estado": "Entregada"})
    pendiente_datos  = await db.inventario_motos.count_documents({"estado": "Pendiente datos"})
    anuladas         = await db.inventario_motos.count_documents({"estado": "Anulada"})
    pipeline = [{"$group": {"_id": None, "total_inversion": {"$sum": "$total"}, "total_costo": {"$sum": "$costo"}}}]
    agg = await db.inventario_motos.aggregate(pipeline).to_list(1)
    totals = agg[0] if agg else {"total_inversion": 0, "total_costo": 0}
    return {
        "total": total,
        "disponibles": disponibles,
        "vendidas": vendidas,
        "entregadas": entregadas,
        "pendiente_datos": pendiente_datos,
        "anuladas": anuladas,
        "total_inversion": totals.get("total_inversion", 0),
        "total_costo": totals.get("total_costo", 0),
        "ultima_actualizacion": datetime.now(timezone.utc).isoformat(),
    }


@router.get("/auditoria")
async def auditoria_inventario(current_user=Depends(get_current_user)):
    """Audita el inventario: detecta inconsistencias entre motos y loanbooks."""
    total = await db.inventario_motos.count_documents({})
    disponibles = await db.inventario_motos.count_documents({"estado": "Disponible"})
    vendidas_entregadas = await db.inventario_motos.count_documents(
        {"estado": {"$in": ["Vendida", "Entregada"]}}
    )
    anuladas = await db.inventario_motos.count_documents({"estado": "Anulada"})
    loanbooks_activos = await db.loanbook.count_documents(
        {"estado": {"$in": ["activo", "mora", "pendiente_entrega"]}}
    )

    inconsistencias = []

    # 1. Motos fantasma (sin chasis válido o con marca inesperada)
    marcas_validas = ["TVS", "Auteco", "Kawasaki", "Honda", "Yamaha", "Bajaj", "Hero", "AKT"]
    fantasmas_cursor = db.inventario_motos.find({
        "$or": [
            {"chasis": None},
            {"chasis": ""},
            {"chasis": {"$regex": "^PENDIENTE-"}},
            {"marca": {"$nin": marcas_validas}},
        ]
    }, {"_id": 0, "id": 1, "marca": 1, "modelo": 1, "chasis": 1, "estado": 1, "factura_compra": 1})
    async for m in fantasmas_cursor:
        inconsistencias.append({
            "tipo": "moto_sin_chasis_valido",
            "moto_id": m.get("id"),
            "marca": m.get("marca"),
            "modelo": m.get("modelo"),
            "chasis": m.get("chasis"),
            "accion_sugerida": "Revisar o eliminar si no tiene factura de compra",
        })

    # 2. Motos vendidas/entregadas sin loanbook
    vend_cursor = db.inventario_motos.find(
        {"estado": {"$in": ["Vendida", "Entregada"]}},
        {"_id": 0, "id": 1, "chasis": 1, "propietario": 1, "loanbook_id": 1}
    )
    async for m in vend_cursor:
        lb = None
        if m.get("loanbook_id"):
            lb = await db.loanbook.find_one({"id": m["loanbook_id"]}, {"_id": 0, "id": 1})
        if not lb and m.get("chasis"):
            lb = await db.loanbook.find_one({"moto_chasis": m["chasis"]}, {"_id": 0, "id": 1})
        if not lb:
            inconsistencias.append({
                "tipo": "moto_vendida_sin_loanbook",
                "moto_id": m.get("id"),
                "chasis": m.get("chasis"),
                "propietario": m.get("propietario"),
                "accion_sugerida": "Verificar manualmente — puede ser venta sin crédito",
            })

    # 3. Loanbooks activos sin moto en inventario
    lb_cursor = db.loanbook.find(
        {"estado": {"$in": ["activo", "mora", "pendiente_entrega"]}},
        {"_id": 0, "id": 1, "codigo": 1, "cliente_nombre": 1, "moto_chasis": 1}
    )
    async for lb in lb_cursor:
        chasis = lb.get("moto_chasis")
        if not chasis:
            inconsistencias.append({
                "tipo": "loanbook_sin_vin",
                "loanbook_codigo": lb.get("codigo"),
                "cliente": lb.get("cliente_nombre"),
                "accion_sugerida": "Asignar VIN desde facturas de venta Alegra",
            })
            continue
        moto = await db.inventario_motos.find_one({"chasis": chasis}, {"_id": 0, "id": 1, "estado": 1})
        if not moto:
            inconsistencias.append({
                "tipo": "loanbook_moto_no_en_inventario",
                "loanbook_codigo": lb.get("codigo"),
                "cliente": lb.get("cliente_nombre"),
                "chasis": chasis,
                "accion_sugerida": "Registrar moto faltante en inventario",
            })

    cuadra = vendidas_entregadas == loanbooks_activos
    return {
        "total_motos": total,
        "disponibles": disponibles,
        "vendidas_entregadas": vendidas_entregadas,
        "anuladas": anuladas,
        "loanbooks_activos": loanbooks_activos,
        "cuadra": cuadra,
        "inconsistencias": inconsistencias,
        "timestamp": datetime.now(timezone.utc).isoformat(),
    }


@router.delete("/motos/{moto_id}")
async def eliminar_moto(
    moto_id: str,
    body: dict = None,
    current_user=Depends(require_admin),
):
    """Elimina una moto del inventario con verificación y log de evento."""
    body = body or {}
    moto = await db.inventario_motos.find_one({"id": moto_id}, {"_id": 0})
    if not moto:
        raise HTTPException(status_code=404, detail="Moto no encontrada")

    # Verify no active loanbook
    chasis = moto.get("chasis", "")
    lb = None
    if chasis and not chasis.startswith("PENDIENTE-"):
        lb = await db.loanbook.find_one(
            {"$or": [{"moto_chasis": chasis}, {"moto_id": moto_id}]},
            {"_id": 0, "codigo": 1, "cliente_nombre": 1, "estado": 1}
        )
        if lb and lb.get("estado") in ("activo", "mora", "pendiente_entrega"):
            raise HTTPException(
                status_code=400,
                detail=f"No se puede eliminar: la moto tiene loanbook activo {lb.get('codigo')} — {lb.get('cliente_nombre')}",
            )

    await db.inventario_motos.delete_one({"id": moto_id})
    await db.roddos_events.insert_one({
        "event_type": "inventario.moto.eliminada",
        "motivo": body.get("motivo", "Eliminada por administrador"),
        "registro_eliminado": moto,
        "timestamp": datetime.now(timezone.utc).isoformat(),
        "ejecutado_por": current_user.get("email", "admin"),
    })
    return {"ok": True, "mensaje": f"Moto {moto.get('marca','')} {moto.get('modelo','')} chasis '{chasis}' eliminada del inventario."}


@router.post("/asignar-chasis")
async def asignar_chasis_loanbook(
    body: dict,
    current_user=Depends(require_admin),
):
    """Asigna VIN y motor a un loanbook y actualiza el estado de la moto."""
    loanbook_codigo = body.get("loanbook_codigo")
    chasis = body.get("chasis", "").strip()
    motor = body.get("motor", "").strip()

    if not loanbook_codigo or not chasis:
        raise HTTPException(status_code=400, detail="Se requieren loanbook_codigo y chasis")

    lb = await db.loanbook.find_one({"codigo": loanbook_codigo}, {"_id": 0, "id": 1, "estado": 1, "cliente_nombre": 1})
    if not lb:
        raise HTTPException(status_code=404, detail=f"Loanbook {loanbook_codigo} no encontrado")

    moto = await db.inventario_motos.find_one({"chasis": chasis}, {"_id": 0, "id": 1, "estado": 1})
    if not moto:
        raise HTTPException(status_code=404, detail=f"Moto con chasis {chasis} no encontrada en inventario")

    now = datetime.now(timezone.utc).isoformat()
    lb_estado = lb.get("estado", "activo")
    moto_nuevo_estado = "Entregada" if lb_estado in ("activo", "mora") else "Vendida"

    await db.loanbook.update_one(
        {"codigo": loanbook_codigo},
        {"$set": {"moto_chasis": chasis, "motor": motor, "updated_at": now}},
    )
    await db.inventario_motos.update_one(
        {"chasis": chasis},
        {"$set": {
            "estado": moto_nuevo_estado,
            "propietario": lb.get("cliente_nombre", ""),
            "loanbook_id": lb.get("id", ""),
            "loanbook_codigo": loanbook_codigo,
            "updated_at": now,
        }},
    )
    return {
        "ok": True,
        "loanbook": loanbook_codigo,
        "chasis": chasis,
        "motor": motor,
        "moto_estado": moto_nuevo_estado,
    }


@router.put("/motos/{moto_id}")
async def update_moto(moto_id: str, body: dict, current_user=Depends(get_current_user)):
    body.pop("_id", None)
    body.pop("id", None)
    result = await db.inventario_motos.update_one({"id": moto_id}, {"$set": body})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Moto no encontrada")
    updated = await db.inventario_motos.find_one({"id": moto_id}, {"_id": 0})
    return updated


@router.post("/motos/{moto_id}/register-alegra")
async def register_in_alegra(moto_id: str, current_user=Depends(get_current_user)):
    moto = await db.inventario_motos.find_one({"id": moto_id}, {"_id": 0})
    if not moto:
        raise HTTPException(status_code=404, detail="Moto no encontrada")
    if moto.get("alegra_item_id"):
        raise HTTPException(status_code=400, detail="Esta moto ya está registrada en Alegra")
    service = AlegraService(db)
    result = await register_moto_in_alegra(moto, service)
    alegra_id = result.get("id")
    await db.inventario_motos.update_one({"id": moto_id}, {"$set": {"alegra_item_id": alegra_id}})
    await log_action(current_user, f"/inventario/motos/{moto_id}/register-alegra", "POST", {"alegra_id": alegra_id})
    return {"alegra_item_id": alegra_id, "result": result}


class VentaMotoRequest(BaseModel):
    cliente_id: str
    cliente_nombre: str
    precio_venta: float
    tipo_pago: str = "contado"
    cuotas: int = 1
    valor_cuota: Optional[float] = None
    include_iva: bool = True
    ipoc_pct: float = 8.0


@router.post("/motos/{moto_id}/vender")
async def vender_moto(moto_id: str, req: VentaMotoRequest, current_user=Depends(get_current_user)):
    moto = await db.inventario_motos.find_one({"id": moto_id}, {"_id": 0})
    if not moto:
        raise HTTPException(status_code=404, detail="Moto no encontrada")
    if moto.get("estado") != "Disponible":
        raise HTTPException(status_code=400, detail=f"Moto no disponible (estado: {moto.get('estado')})")

    service = AlegraService(db)

    items = [{
        "description": f"{moto.get('marca')} {moto.get('version')} — Chasis: {moto.get('chasis')} Motor: {moto.get('motor')}",
        "quantity": 1,
        "price": req.precio_venta,
        "account": {"id": "4105"},
    }]
    if req.include_iva:
        items[0]["tax"] = [{"percentage": 19}]

    invoice_payload = {
        "date": datetime.now(timezone.utc).strftime("%Y-%m-%d"),
        "dueDate": datetime.now(timezone.utc).strftime("%Y-%m-%d"),
        "client": {"id": req.cliente_id},
        "items": items,
        "observations": f"Venta moto {moto.get('marca')} {moto.get('version')} — {req.tipo_pago}",
    }

    result = await service.request("invoices", "POST", invoice_payload)
    invoice_id = result.get("id")
    invoice_number = (
        result.get("numberTemplate", {}).get("fullNumber")
        or result.get("number")
        or str(invoice_id)
    )

    sale_data = {
        "estado": "Vendida",
        "cliente_id": req.cliente_id,
        "cliente_nombre": req.cliente_nombre,
        "precio_venta": req.precio_venta,
        "tipo_pago": req.tipo_pago,
        "cuotas": req.cuotas,
        "valor_cuota": req.valor_cuota,
        "factura_id": invoice_id,
        "factura_numero": invoice_number,
        "fecha_venta": datetime.now(timezone.utc).isoformat(),
        "vendido_por": current_user.get("email"),
    }
    await db.inventario_motos.update_one({"id": moto_id}, {"$set": sale_data})

    if moto.get("alegra_item_id"):
        try:
            await service.request(f"items/{moto['alegra_item_id']}", "PUT", {"status": "inactive"})
        except Exception:
            pass

    await log_action(
        current_user, f"/inventario/motos/{moto_id}/vender", "POST",
        {"invoice_number": invoice_number, "cliente": req.cliente_nombre},
    )

    return {
        "success": True,
        "invoice_id": invoice_id,
        "invoice_number": invoice_number,
        "message": f"Moto vendida — Factura {invoice_number} creada en Alegra",
        "moto": {**moto, **sale_data},
    }


# ── BUILD 12: Carga inicial de costos de inventario ───────────────────────────

@router.get("/plantilla-costos")
async def descargar_plantilla_costos(current_user=Depends(require_admin)):
    """Genera y descarga plantilla Excel para carga masiva de costos de inventario."""
    import io
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

    wb  = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Costos"

    hdr_fill = PatternFill("solid", fgColor="0F2A5C")
    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    ex_fill  = PatternFill("solid", fgColor="E8F0FE")
    thin     = Side(style="thin", color="CCCCCC")
    brd      = Border(left=thin, right=thin, top=thin, bottom=thin)

    COLS = ["Chasis", "Referencia", "Color", "Costo_Unitario",
            "IVA_Unitario", "IPOC_Unitario", "Proveedor", "Fecha_Compra"]
    for i, col in enumerate(COLS, 1):
        c = ws1.cell(1, i, col)
        c.fill = hdr_fill; c.font = hdr_font; c.border = brd
        c.alignment = Alignment(horizontal="center", vertical="center")

    # Example row
    EXAMPLE = ["TEST-EJEMPLO", "Honda CB190R", "Rojo", 7058824, 1341176, 0, "Auteco", "2026-01-15"]
    for i, val in enumerate(EXAMPLE, 1):
        c = ws1.cell(2, i, val)
        c.fill = ex_fill; c.border = brd
        c.alignment = Alignment(horizontal="center" if i > 3 else "left")

    widths = [22, 20, 14, 16, 14, 14, 16, 16]
    for i, w in enumerate(widths, 1):
        ws1.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # Hoja 2 Instrucciones
    ws2 = wb.create_sheet("Instrucciones")
    ws2["A1"].value = "INSTRUCCIONES — Carga de Costos de Inventario RODDOS"
    ws2["A1"].font = Font(bold=True, color="0F2A5C", size=13)
    ws2.merge_cells("A1:D1")

    fields = [
        ("Chasis",         "Número de chasis exacto como aparece en MongoDB",              "3HLTV18J86M700001", "✅"),
        ("Referencia",     "Modelo de la moto",                                            "Honda CB190R",       "✅"),
        ("Color",          "Color de la unidad",                                           "Rojo",               "✅"),
        ("Costo_Unitario", "Precio de compra sin IVA ni IPOC (pesos, sin puntos)",         "7058824",            "✅"),
        ("IVA_Unitario",   "IVA pagado por unidad (19% del costo). 0 si no aplica",       "1341176",            "✅"),
        ("IPOC_Unitario",  "IPOC pagado. 0 si no aplica",                                  "0",                  "✅"),
        ("Proveedor",      "Nombre del proveedor (Auteco, otro)",                          "Auteco",             "✅"),
        ("Fecha_Compra",   "Fecha de compra en formato YYYY-MM-DD",                        "2026-01-15",         "✅"),
    ]
    hdr_li_fill = PatternFill("solid", fgColor="D9E1F2")
    for col, h in enumerate(["Campo", "Descripción", "Ejemplo", "Req."], 1):
        c = ws2.cell(3, col, h)
        c.font = Font(bold=True); c.fill = hdr_li_fill; c.border = brd

    for r, (f, d, e, ob) in enumerate(fields, 4):
        ws2.cell(r, 1, f).font = Font(bold=True)
        ws2.cell(r, 2, d); ws2.cell(r, 3, e); ws2.cell(r, 4, ob)
        for col in range(1, 5): ws2.cell(r, col).border = brd

    notas_start = len(fields) + 6
    for i, nota in enumerate([
        "NOTAS IMPORTANTES:",
        "• No modificar los encabezados de la columna Chasis",
        "• La fila de ejemplo (TEST-EJEMPLO) es ignorada automáticamente",
        "• Solo se actualizan motos que ya existen en el sistema",
        "• Guardar como .xlsx antes de subir",
    ]):
        c = ws2.cell(notas_start + i, 1, nota)
        if i == 0: c.font = Font(bold=True, color="0F2A5C")
        ws2.merge_cells(f"A{notas_start+i}:D{notas_start+i}")

    ws2.column_dimensions["A"].width = 20
    ws2.column_dimensions["B"].width = 52
    ws2.column_dimensions["C"].width = 20
    ws2.column_dimensions["D"].width = 10

    from fastapi.responses import StreamingResponse
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="RODDOS_Plantilla_Costos.xlsx"'},
    )


@router.post("/cargar-costos/preview")
async def preview_cargar_costos(
    file: UploadFile = File(...),
    current_user=Depends(require_admin),
):
    """Preview: cruza chasis del Excel con inventario_motos. No guarda aún."""
    import io as _io, openpyxl as _opx
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="El archivo debe ser .xlsx o .xls")

    contents = await file.read()
    try:
        wb = _opx.load_workbook(_io.BytesIO(contents), data_only=True)
        ws = wb.active
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Error al leer Excel: {e}")

    headers = [str(c.value or "").strip().lower() for c in ws[1]]
    def _col(name):
        for aliases in [[name], [name.replace("_"," ")]]:
            for a in aliases:
                if a in headers:
                    return headers.index(a)
        return None

    idx_chasis = _col("chasis")
    if idx_chasis is None:
        return {"ok": False, "error": "No se encontró la columna 'Chasis' en el archivo."}

    idx_costo  = _col("costo_unitario")
    idx_iva    = _col("iva_unitario")
    idx_ipoc   = _col("ipoc_unitario")
    idx_prov   = _col("proveedor")
    idx_fecha  = _col("fecha_compra")

    actualizadas, no_encontradas, errores = [], [], []
    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        chasis = str(row[idx_chasis] or "").strip() if idx_chasis is not None and idx_chasis < len(row) else ""
        if not chasis or chasis.upper() == "TEST-EJEMPLO":
            continue
        moto = await db.inventario_motos.find_one({"chasis": chasis}, {"_id": 0, "referencia": 1, "id": 1})
        if not moto:
            no_encontradas.append({"fila": row_num, "chasis": chasis})
            continue

        def _num(idx):
            try: return float(row[idx]) if idx is not None and idx < len(row) and row[idx] else 0.0
            except: return 0.0

        actualizadas.append({
            "fila":         row_num,
            "chasis":       chasis,
            "moto_id":      moto.get("id"),
            "referencia":   moto.get("referencia"),
            "precio_compra": _num(idx_costo),
            "iva_compra":    _num(idx_iva),
            "ipoc_compra":   _num(idx_ipoc),
            "proveedor_compra": str(row[idx_prov] or "").strip() if idx_prov is not None else "",
            "fecha_compra":  str(row[idx_fecha] or "")[:10] if idx_fecha is not None else "",
        })

    return {
        "ok":             True,
        "total_filas":    row_num - 1,
        "actualizadas":   actualizadas,
        "no_encontradas": no_encontradas,
        "resumen":        f"{len(actualizadas)} motos a actualizar, {len(no_encontradas)} chasis no encontrados.",
    }


@router.post("/cargar-costos/confirmar")
async def confirmar_cargar_costos(
    payload: dict,
    current_user=Depends(require_admin),
):
    """Guarda los costos confirmados en inventario_motos y emite evento."""
    actualizadas = payload.get("actualizadas", [])
    if not actualizadas:
        raise HTTPException(status_code=400, detail="No hay datos para guardar.")

    ok = 0
    for item in actualizadas:
        moto_id = item.get("moto_id")
        if not moto_id:
            continue
        update_data = {
            "precio_compra":    item.get("precio_compra", 0),
            "iva_compra":       item.get("iva_compra", 0),
            "ipoc_compra":      item.get("ipoc_compra", 0),
            "proveedor_compra": item.get("proveedor_compra", ""),
            "fecha_compra":     item.get("fecha_compra", ""),
            "costos_cargados":  True,
            "updated_at":       datetime.now(timezone.utc).isoformat(),
        }
        result = await db.inventario_motos.update_one({"id": moto_id}, {"$set": update_data})
        if result.modified_count:
            ok += 1

    # Mark initial load as done (flag para ocultar el botón)
    await db.system_config.update_one(
        {"key": "inventario_costos_cargados"},
        {"$set": {"key": "inventario_costos_cargados", "value": True,
                  "fecha": datetime.now(timezone.utc).isoformat(),
                  "usuario": current_user.get("email")}},
        upsert=True,
    )

    # Emit event
    await db.roddos_events.insert_one({
        "event_type":   "inventario.costos.actualizados",
        "motos_ok":     ok,
        "total":        len(actualizadas),
        "usuario":      current_user.get("email"),
        "created_at":   datetime.now(timezone.utc).isoformat(),
    })

    return {
        "ok":      True,
        "guardadas": ok,
        "mensaje": f"✅ Costos actualizados: {ok} motos. El P&L ahora muestra el costo real de ventas.",
    }


@router.get("/costos-cargados")
async def check_costos_cargados(current_user=Depends(require_admin)):
    """Retorna si la carga inicial de costos ya fue realizada."""
    doc = await db.system_config.find_one({"key": "inventario_costos_cargados"}, {"_id": 0})
    return {"cargados": bool(doc and doc.get("value"))}


@router.patch("/motos/{moto_id}/costo")
async def update_costo_moto(
    moto_id: str,
    body: dict,
    current_user=Depends(require_admin),
):
    """Actualiza costo de compra de una moto individual (solo admin)."""
    allowed = {"precio_compra", "iva_compra", "ipoc_compra", "proveedor_compra", "fecha_compra"}
    update  = {k: v for k, v in body.items() if k in allowed}
    if not update:
        raise HTTPException(status_code=400, detail="No hay campos de costo para actualizar.")

    update["costos_cargados"] = True
    update["updated_at"]      = datetime.now(timezone.utc).isoformat()

    result = await db.inventario_motos.update_one({"id": moto_id}, {"$set": update})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Moto no encontrada")

    moto = await db.inventario_motos.find_one({"id": moto_id}, {"_id": 0})
    return {"ok": True, "moto": moto}
