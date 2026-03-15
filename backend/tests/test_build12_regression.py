"""BUILD 12 regression tests: P&L, indicadores, inventario costos, AI chat"""
import pytest
import requests
import os
import io

BASE_URL = os.environ.get("REACT_APP_BACKEND_URL", "").rstrip("/")
EMAIL = "contabilidad@roddos.com"
PASSWORD = "Admin@RODDOS2025!"


@pytest.fixture(scope="module")
def token():
    r = requests.post(f"{BASE_URL}/api/auth/login", json={"email": EMAIL, "password": PASSWORD})
    if r.status_code != 200:
        pytest.skip(f"Auth failed: {r.status_code} {r.text}")
    return r.json().get("access_token") or r.json().get("token")


@pytest.fixture(scope="module")
def auth(token):
    return {"Authorization": f"Bearer {token}"}


# ── Test 1: GET /api/cfo/indicadores ─────────────────────────────────────────
class TestIndicadores:
    def test_indicadores_returns_200(self, auth):
        r = requests.get(f"{BASE_URL}/api/cfo/indicadores", headers=auth)
        assert r.status_code == 200, f"Expected 200, got {r.status_code}: {r.text}"

    def test_indicadores_creditos_activos(self, auth):
        r = requests.get(f"{BASE_URL}/api/cfo/indicadores", headers=auth)
        data = r.json()
        activos = data.get("creditos_activos", -1)
        print(f"creditos_activos={activos}")
        assert activos == 10, f"Expected 10 activos, got {activos}"

    def test_indicadores_creditos_minimos(self, auth):
        r = requests.get(f"{BASE_URL}/api/cfo/indicadores", headers=auth)
        data = r.json()
        minimos = data.get("creditos_minimos", -1)
        print(f"creditos_minimos={minimos}")
        assert minimos == 45, f"Expected 45 creditos_minimos, got {minimos}"

    def test_indicadores_recaudo(self, auth):
        r = requests.get(f"{BASE_URL}/api/cfo/indicadores", headers=auth)
        data = r.json()
        recaudo = data.get("recaudo_semanal_base", 0)
        print(f"recaudo_semanal_base={recaudo}")
        # recaudo should be around 1,659,400
        assert recaudo > 0, f"recaudo_semanal_base should be > 0, got {recaudo}"

    def test_indicadores_margen_semanal(self, auth):
        r = requests.get(f"{BASE_URL}/api/cfo/indicadores", headers=auth)
        data = r.json()
        margen = data.get("margen_semanal", 0)
        gastos = data.get("gastos_fijos_semanales", 0)
        recaudo = data.get("recaudo_semanal_base", 0)
        print(f"margen_semanal={margen}, gastos={gastos}, recaudo={recaudo}")
        # Margen = recaudo - gastos (should be ~-5,840,600)
        if gastos > 0 and recaudo > 0:
            expected = recaudo - gastos
            assert abs(margen - expected) < 1000, f"margen_semanal={margen} should be recaudo-gastos={expected}"


# ── Test 2: GET /api/cfo/estado-resultados ────────────────────────────────────
class TestEstadoResultados:
    def test_pl_returns_200(self, auth):
        r = requests.get(f"{BASE_URL}/api/cfo/estado-resultados?periodo=2026-03", headers=auth)
        assert r.status_code == 200, f"Expected 200, got {r.status_code}: {r.text}"

    def test_pl_has_required_fields(self, auth):
        r = requests.get(f"{BASE_URL}/api/cfo/estado-resultados?periodo=2026-03", headers=auth)
        data = r.json()
        assert "modo" in data, "Missing field: modo"
        assert "ingresos" in data, "Missing field: ingresos"
        assert "total" in data["ingresos"], "Missing field: ingresos.total"
        assert "utilidad_neta" in data, "Missing field: utilidad_neta"
        assert "comparativo" in data, "Missing field: comparativo"

    def test_pl_modo_valid(self, auth):
        r = requests.get(f"{BASE_URL}/api/cfo/estado-resultados?periodo=2026-03", headers=auth)
        data = r.json()
        modo = data.get("modo")
        print(f"P&L modo={modo}, ingresos_total={data.get('ingresos',{}).get('total')}, utilidad_neta={data.get('utilidad_neta')}")
        assert modo in ("parcial", "completo"), f"modo should be 'parcial' or 'completo', got {modo}"


# ── Test 3: GET /api/inventario/plantilla-costos ───────────────────────────────
class TestPlantillaCostos:
    def test_plantilla_returns_excel(self, auth):
        r = requests.get(f"{BASE_URL}/api/inventario/plantilla-costos", headers=auth)
        assert r.status_code == 200, f"Expected 200, got {r.status_code}: {r.text}"
        ct = r.headers.get("content-type", "")
        assert "spreadsheetml" in ct or "excel" in ct or "octet-stream" in ct, \
            f"Expected Excel content-type, got {ct}"

    def test_plantilla_has_headers(self, auth):
        import openpyxl
        r = requests.get(f"{BASE_URL}/api/inventario/plantilla-costos", headers=auth)
        wb = openpyxl.load_workbook(io.BytesIO(r.content))
        ws = wb.active
        headers = [str(c.value or "").strip() for c in ws[1]]
        print(f"Plantilla costos headers: {headers}")
        # Check at least some expected headers
        assert any("referencia" in h.lower() or "placa" in h.lower() or "vin" in h.lower() for h in headers), \
            f"Expected identifier column in headers: {headers}"
        assert any("costo" in h.lower() or "precio" in h.lower() for h in headers), \
            f"Expected cost column in headers: {headers}"


# ── Test 4: POST /api/inventario/cargar-costos/preview ───────────────────────
class TestCargarCostosPreview:
    def test_preview_returns_fields(self, auth):
        """Create minimal xlsx and send to preview endpoint"""
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        # Get real headers first
        r_tmpl = requests.get(f"{BASE_URL}/api/inventario/plantilla-costos", headers=auth)
        if r_tmpl.status_code != 200:
            pytest.skip("Could not get plantilla-costos template")
        
        wb_tmpl = openpyxl.load_workbook(io.BytesIO(r_tmpl.content))
        ws_tmpl = wb_tmpl.active
        headers = [str(c.value or "").strip() for c in ws_tmpl[1]]
        print(f"Template headers for preview test: {headers}")
        
        # Write headers to new workbook
        for col, h in enumerate(headers, 1):
            ws.cell(1, col, h)
        
        # Write one sample row (referencing columns by name)
        header_lower = [h.lower() for h in headers]
        row_vals = [""] * len(headers)
        for i, h in enumerate(header_lower):
            if "referencia" in h or "placa" in h or "vin" in h:
                row_vals[i] = "TEST-VIN-001"
            if "costo" in h or "precio" in h:
                row_vals[i] = 5000000
        for col, v in enumerate(row_vals, 1):
            ws.cell(2, col, v)
        
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        
        files = {"file": ("test_costos.xlsx", buf.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        r = requests.post(f"{BASE_URL}/api/inventario/cargar-costos/preview", headers=auth, files=files)
        print(f"preview response: {r.status_code} {r.text[:200]}")
        assert r.status_code == 200, f"Expected 200, got {r.status_code}: {r.text}"
        data = r.json()
        # Should have actualizadas or no_encontradas
        assert "actualizadas" in data or "no_encontradas" in data or "ok" in data, \
            f"Missing expected fields in response: {data}"


# ── Test 5: GET /api/cfo/estado-resultados/pdf ────────────────────────────────
class TestEstadoResultadosPDF:
    def test_pdf_returns_200(self, auth):
        r = requests.get(f"{BASE_URL}/api/cfo/estado-resultados/pdf?periodo=2026-03", headers=auth)
        assert r.status_code == 200, f"Expected 200, got {r.status_code}: {r.text[:200]}"
        ct = r.headers.get("content-type", "")
        assert "pdf" in ct, f"Expected PDF content-type, got {ct}"


# ── Test 6: GET /api/cfo/estado-resultados/excel ─────────────────────────────
class TestEstadoResultadosExcel:
    def test_excel_returns_200(self, auth):
        r = requests.get(f"{BASE_URL}/api/cfo/estado-resultados/excel?periodo=2026-03", headers=auth)
        assert r.status_code == 200, f"Expected 200, got {r.status_code}: {r.text[:200]}"
        ct = r.headers.get("content-type", "")
        assert "spreadsheetml" in ct or "excel" in ct or "octet-stream" in ct, \
            f"Expected Excel content-type, got {ct}"
