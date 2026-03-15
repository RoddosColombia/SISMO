"""Tests for CFO deudas plantilla download and upload features."""
import pytest
import requests
import os
import io

BASE_URL = os.environ.get("REACT_APP_BACKEND_URL", "").rstrip("/")

EMAIL = "contabilidad@roddos.com"
PASSWORD = "Admin@RODDOS2025!"


@pytest.fixture(scope="module")
def token():
    resp = requests.post(f"{BASE_URL}/api/auth/login", json={"email": EMAIL, "password": PASSWORD})
    assert resp.status_code == 200, f"Login failed: {resp.text}"
    return resp.json()["token"]


@pytest.fixture(scope="module")
def auth_headers(token):
    return {"Authorization": f"Bearer {token}"}


# ── Plantilla download tests ──────────────────────────────────────────────────

def test_descargar_plantilla_status_200(auth_headers):
    resp = requests.get(f"{BASE_URL}/api/cfo/deudas/plantilla", headers=auth_headers)
    assert resp.status_code == 200, f"Expected 200, got {resp.status_code}: {resp.text[:200]}"


def test_descargar_plantilla_content_type(auth_headers):
    resp = requests.get(f"{BASE_URL}/api/cfo/deudas/plantilla", headers=auth_headers)
    ct = resp.headers.get("content-type", "")
    assert "spreadsheetml" in ct or "openxmlformats" in ct or "excel" in ct, f"Unexpected content-type: {ct}"


def test_descargar_plantilla_size_gt_zero(auth_headers):
    resp = requests.get(f"{BASE_URL}/api/cfo/deudas/plantilla", headers=auth_headers)
    assert len(resp.content) > 0, "Empty file returned"


def test_descargar_plantilla_dos_hojas(auth_headers):
    import openpyxl
    resp = requests.get(f"{BASE_URL}/api/cfo/deudas/plantilla", headers=auth_headers)
    wb = openpyxl.load_workbook(io.BytesIO(resp.content))
    assert "Deudas" in wb.sheetnames, f"Missing 'Deudas' sheet. Got: {wb.sheetnames}"
    assert "Instrucciones" in wb.sheetnames, f"Missing 'Instrucciones' sheet. Got: {wb.sheetnames}"


def test_descargar_plantilla_headers_correctos(auth_headers):
    import openpyxl
    resp = requests.get(f"{BASE_URL}/api/cfo/deudas/plantilla", headers=auth_headers)
    wb = openpyxl.load_workbook(io.BytesIO(resp.content))
    ws = wb["Deudas"]
    expected = ["Acreedor", "Descripcion", "Monto_Total", "Monto_Pagado",
                "Tasa_Mensual_Pct", "Fecha_Vencimiento", "Tipo", "Prioridad"]
    actual = [ws.cell(row=1, column=i+1).value for i in range(8)]
    assert actual == expected, f"Headers mismatch: {actual}"


def test_descargar_plantilla_fila_ejemplo_auteco(auth_headers):
    import openpyxl
    resp = requests.get(f"{BASE_URL}/api/cfo/deudas/plantilla", headers=auth_headers)
    wb = openpyxl.load_workbook(io.BytesIO(resp.content))
    ws = wb["Deudas"]
    acreedor_fila2 = ws.cell(row=2, column=1).value
    assert acreedor_fila2 == "Auteco Kawasaki", f"Expected 'Auteco Kawasaki' in row 2, got: {acreedor_fila2}"


def test_descargar_plantilla_instrucciones_titulo(auth_headers):
    import openpyxl
    resp = requests.get(f"{BASE_URL}/api/cfo/deudas/plantilla", headers=auth_headers)
    wb = openpyxl.load_workbook(io.BytesIO(resp.content))
    ws = wb["Instrucciones"]
    title = ws.cell(row=1, column=1).value
    assert title and "INSTRUCCIONES" in str(title).upper(), f"No title found in Instrucciones: {title}"
    # Check field table headers
    col_a = ws.cell(row=3, column=1).value
    assert col_a and "campo" in str(col_a).lower(), f"Expected 'Campo' in row 3, got: {col_a}"


# ── Cargar tests ──────────────────────────────────────────────────────────────

def _make_excel_only_example():
    """Excel with only the example row (Auteco Kawasaki)."""
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Deudas"
    headers = ["Acreedor", "Descripcion", "Monto_Total", "Monto_Pagado",
               "Tasa_Mensual_Pct", "Fecha_Vencimiento", "Tipo", "Prioridad"]
    ws.append(headers)
    ws.append(["Auteco Kawasaki", "Financiación inventario motos", 45000000, 5000000, 1.5, "2026-06-30", "productiva", 1])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def _make_excel_with_real_rows():
    """Excel with example row + 2 real deudas. One with $ and dots in monto."""
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Deudas"
    headers = ["Acreedor", "Descripcion", "Monto_Total", "Monto_Pagado",
               "Tasa_Mensual_Pct", "Fecha_Vencimiento", "Tipo", "Prioridad"]
    ws.append(headers)
    ws.append(["Auteco Kawasaki", "Financiación inventario motos", 45000000, 5000000, 1.5, "2026-06-30", "productiva", 1])
    ws.append(["Bancolombia", "Crédito rotativo", "$3.500.000", 0, 2.0, "2026-12-31", "productiva", 2])
    ws.append(["Davivienda", "Leasing vehiculo", 20000000, 5000000, 1.8, "2027-03-31", "productiva", 3])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def _make_excel_wrong_columns():
    """Excel with wrong column names."""
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["NombreEmpresa", "CantidadPesos", "FechaLimite"])
    ws.append(["Test SA", 1000000, "2026-01-01"])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def test_cargar_solo_fila_ejemplo_retorna_ok_false(auth_headers):
    data = _make_excel_only_example()
    resp = requests.post(
        f"{BASE_URL}/api/cfo/deudas/cargar",
        headers=auth_headers,
        files={"file": ("test.xlsx", io.BytesIO(data), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert resp.status_code == 200, f"Expected 200, got {resp.status_code}: {resp.text}"
    body = resp.json()
    assert body.get("ok") is False, f"Expected ok=false, got: {body}"
    msg = str(body.get("error", "") + body.get("message", "") + body.get("detalle", "")).lower()
    assert "fila" in msg or "dato" in msg or "no tiene" in msg, f"Missing 'no tiene filas con datos' msg: {body}"


def test_cargar_con_filas_reales_retorna_ok_true(auth_headers):
    data = _make_excel_with_real_rows()
    resp = requests.post(
        f"{BASE_URL}/api/cfo/deudas/cargar",
        headers=auth_headers,
        files={"file": ("test.xlsx", io.BytesIO(data), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert resp.status_code == 200, f"Expected 200, got {resp.status_code}: {resp.text}"
    body = resp.json()
    assert body.get("ok") is True, f"Expected ok=true, got: {body}"
    deudas = body.get("deudas", [])
    assert len(deudas) == 2, f"Expected 2 deudas (example skipped), got: {len(deudas)}. Body: {body}"


def test_cargar_monto_con_signo_pesos_y_puntos(auth_headers):
    """'$3.500.000' should parse correctly to 3500000."""
    data = _make_excel_with_real_rows()
    resp = requests.post(
        f"{BASE_URL}/api/cfo/deudas/cargar",
        headers=auth_headers,
        files={"file": ("test.xlsx", io.BytesIO(data), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    body = resp.json()
    deudas = body.get("deudas", [])
    bancolombia = next((d for d in deudas if "bancolombia" in str(d.get("acreedor", "")).lower()), None)
    assert bancolombia is not None, "Bancolombia row not found"
    assert bancolombia["monto_total"] == 3500000.0, f"Expected 3500000, got: {bancolombia['monto_total']}"


def test_cargar_columnas_incorrectas_retorna_ok_false(auth_headers):
    data = _make_excel_wrong_columns()
    resp = requests.post(
        f"{BASE_URL}/api/cfo/deudas/cargar",
        headers=auth_headers,
        files={"file": ("test.xlsx", io.BytesIO(data), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert resp.status_code == 200, f"Expected 200, got {resp.status_code}: {resp.text}"
    body = resp.json()
    assert body.get("ok") is False, f"Expected ok=false, got: {body}"
    full_msg = str(body).lower()
    assert "colum" in full_msg or "faltante" in full_msg or "plantilla" in full_msg, \
        f"Expected missing columns message, got: {body}"
